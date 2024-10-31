# !/usr/bin/python3
# IMPORT MODULE - START
from tkinter import *
from tkinter import messagebox
import time
from time import sleep, gmtime, strftime
from picamera import PiCamera
import cv2
import numpy as np
from tkinter import filedialog
from PIL import ImageTk, Image, ImageDraw, ImageFont
import serial
from functools import partial
import math
from fractions import Fraction
from threading import *
import os
from tkinter import ttk
import awesometkinter as atk
import tkinter.font as font
import openpyxl
import subprocess
import shutil
import RPi.GPIO as GPIO
from ftplib import FTP
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image as Img
from datetime import *
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
import re
import dns.resolver
import socket
import awesometkinter as atk

import board
import busio
import adafruit_ads1x15.ads1115 as ADS
from adafruit_ads1x15.analog_in import AnalogIn

import adafruit_ds1307

max_voltage = 120
divide_voltage = 4.68
# IMPORT MODULE - END

# COLORS DEFINE - START
MAIN_BACKGROUND_COLOR = "grey88"
TASKBAR_BACKGROUND_COLOR = "grey88"
MAINMENU_BUTTON_DISABLE_COLOR = "grey70"
MAINMENU_BUTTON_ENABLE_COLOR = "yellow green"
MAINMENU_BUTTON_TEXT_COLOR = "black"
FRAME_BACKGROUND_COLOR_1 = "grey88"
FRAME_TEXT_COLOR_1 = "black"
FRAME_BACKGROUND_COLOR_2 = "grey88"
FRAME_TEXT_COLOR_2 = "black"
FRAME_BACKGROUND_COLOR_3 = "grey65"        #realtime voltage background
FRAME_TEXT_COLOR_3 = "black"
LABEL_BACKGROUND_COLOR_1 = "deep sky blue"
LABEL_TEXT_COLOR_1 = "white"
LABEL_BACKGROUND_COLOR_2 = "DeepSkyBlue2"
LABEL_TEXT_COLOR_2 = "black"
LABEL_BACKGROUND_COLOR_3 = "grey88"
LABEL_TEXT_COLOR_3 = "red"
LABEL_BACKGROUND_COLOR_4 = "grey88"
LABEL_TEXT_COLOR_4 = "blue"
LABEL_BACKGROUND_COLOR_5 = "grey88"
LABEL_TEXT_COLOR_5 = "green yellow"              #realtime voltage text
BUTTON_BACKGROUND_COLOR_1 = "DeepSkyBlue2"
BUTTON_TEXT_COLOR_1 = "black"
BUTTON_BACKGROUND_COLOR_2 = "grey92"
BUTTON_TEXT_COLOR_2 = "black"
BUTTON_BACKGROUND_COLOR_3 = "light blue"
BUTTON_TEXT_COLOR_3 = "black"
BUTTON_BACKGROUND_COLOR_4 = "red3"
BUTTON_TEXT_COLOR_4 = "black"
BUTTON_BACKGROUND_COLOR_5 = "lavender"
BUTTON_TEXT_COLOR_5 = "black"

# COLORS DEFINE - END


# GLOBAL VARIABLE - START
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 587

result_path = "/home/pi/Desktop/VE100 Result"
path_name = ""
folder_name = ""
folder_name_set = ""
recipient_email = "@gmail.com"
time_capture = 0
count=1
account_active = -1
email_password = ''
email_address = ''
auto_capture_call = 0
automail1_is_on = 0
automail2_is_on = 0

voltage_set = 36
m_set = 10
s_set = 0
auto_capture_timer_set = 2
solve_s = 0
solve_auto_capture = 0
voltage0_set = 80
voltage1_set = 15
voltage2_set = 20
voltage3_set = 25
m0_set=0
s0_set=15
m1_set=1
s1_set=0
m2_set=2
s2_set=0
m3_Set=0
s3_set=30
solve_s0 = 0
solve_s1 = 0
solve_s2 = 0
solve_s3 = 0
stage0_is_running = 1
stage1_is_running = 0
stage2_is_running = 0
stage3_is_running = 0
final_cap = 0
m1_raw = 0
m2_raw = 0
m3_raw = 0
m4_raw = 0
s1_raw = 0
s2_raw = 0
s3_raw = 0
s4_raw = 0

running = 0

# band_crop_x1 = 25
# band_crop_y1 = 203
# band_crop_x2 = 860
# band_crop_y2 = 465

scale_percent_1 = 76
tmp_value_1 = 4
tmp_value_2 = tmp_value_1*scale_percent_1/100
scale_percent_2 = (tmp_value_1*100)/tmp_value_2

x0_limit = 0
y0_limit = 280
x1_limit = 56
y1_limit = 280

band_finder_clicked = 0

preview_only = 0

CURRENT_WARNING = 0.045
# RELAY_PIN = 20
# BLUE_LIGHT_PIN = 21
# POWER_LED_PIN = 16
POWER_LED_PIN = 13
RELAY_PIN = 26
BLUE_LIGHT_PIN = 19
BUZZER_PIN = 6
RUN_LED_PIN = 16
SENSOR_PIN = 20
# GLOBAL VARIABLE - START

# INIT FOLDER - START
if not os.path.exists('/home/pi/Desktop/VE100 Result'):
	f = os.path.join("/home/pi/Desktop", "VE100 Result")
	os.mkdir(f)

if not os.path.exists('/home/pi/VE100'):
	f = os.path.join("/home/pi/VE100")
	os.mkdir(f)

if not os.path.exists('/home/pi/VE100/.account.txt'):
	fw_account = open("/home/pi/VE100/.account.txt", "x")
	fw_account.writelines('0\n')
	fw_account.close()
	account_active = 0
else:
	fr_account = open("/home/pi/VE100/.account.txt","r")
	account_active = int(fr_account.readline())
	email_address = fr_account.readline().strip('\n')
	email_password = fr_account.readline().strip('\n')

if not os.path.exists('/home/pi/VE100/parameters1.txt'):
	fw_para1 = open('/home/pi/VE100/parameters1.txt', 'x')
	fw_para1.writelines('36\n')
	fw_para1.writelines('1000\n')
	fw_para1.writelines('5\n')
	fw_para1.writelines('0\n')
	fw_para1.close()
else:
	fr_para1 = open("/home/pi/VE100/parameters1.txt","r")
	voltage_set = int(fr_para1.readline())
	timer_set = int(fr_para1.readline().strip('\n'))
	auto_capture_timer_set = int(fr_para1.readline().strip('\n'))
	automail1_is_on = int(fr_para1.readline().strip('\n'))

	m_set = round(timer_set/100)
	s_set = timer_set - (m_set*100)

if not os.path.exists('/home/pi/VE100/parameters2.txt'):
	fw_para2 = open('/home/pi/VE100/parameters2.txt', 'x')
	fw_para2.writelines('20\n')
	fw_para2.writelines('0500\n')
	fw_para2.writelines('25\n')
	fw_para2.writelines('1000\n')
	fw_para2.writelines('30\n')
	fw_para2.writelines('1500\n')
	fw_para2.writelines('0\n')
	fw_para2.close()
else:
	fr_para2 = open("/home/pi/VE100/parameters2.txt","r")
	voltage1_set = int(fr_para2.readline())
	timer1_set = int(fr_para2.readline().strip('\n'))
	voltage2_set = int(fr_para2.readline().strip('\n'))
	timer2_set = int(fr_para2.readline().strip('\n'))
	voltage3_set = int(fr_para2.readline().strip('\n'))
	timer3_set = int(fr_para2.readline().strip('\n'))
	automail2_is_on = int(fr_para2.readline().strip('\n'))

	m1_set = timer1_set//100
	s1_set = timer1_set - (m1_set*100)
	m2_set = timer2_set//100
	s2_set = timer2_set - (m2_set*100)
	m3_set = timer3_set//100
	s3_set = timer3_set - (m3_set*100)

if not os.path.exists('/home/pi/VE100/.config.txt'):
	fw_voltage = open('/home/pi/VE100/.config.txt', 'x')
	fw_voltage.writelines('4.68\n')
	fw_voltage.close()
else:
	fr_voltage = open('/home/pi/VE100/.config.txt')
	divide_voltage = float(fr_voltage.readline())
	print("Divide Voltage: ", divide_voltage)

if not os.path.exists('/home/pi/VE100/coordinates.txt'):
	fw_coordinate = open('/home/pi/VE100/coordinates.txt', 'x')
	fw_coordinate.writelines('96\n') # toa do x 18 gieng
	fw_coordinate.writelines('90\n') # toa do x 26 gieng
	fw_coordinate.writelines('230\n') # toa do y chung
	fw_coordinate.writelines('49\n') # khoang cach cac gieng 18
	fw_coordinate.writelines('67\n') # khoang giua 18
	fw_coordinate.writelines('32\n') # khoang cach cac gieng 26
	fw_coordinate.writelines('77\n') # khoang giua 26
	fw_coordinate.writelines('15\n') # font size 18
	fw_coordinate.writelines('15\n') # font size 26
	fw_coordinate.close()
	x_coordinate_17 = 96
	x_coordinate_26 = 90
	y_coordinate = 230
	well_distance_17 = 49
	pace_17 = 67
	well_distance_26 = 32
	pace_26 = 77
	font_size_17 = 15
	font_size_26 = 15
	
else:
	fr_coordinate = open('/home/pi/VE100/coordinates.txt')
	x_coordinate_17 = int(fr_coordinate.readline())
	x_coordinate_26 = int(fr_coordinate.readline().strip('\n'))
	y_coordinate = int(fr_coordinate.readline().strip('\n'))
	well_distance_17 = int(fr_coordinate.readline().strip('\n'))
	pace_17 = int(fr_coordinate.readline().strip('\n'))
	well_distance_26 = int(fr_coordinate.readline().strip('\n'))
	pace_26 = int(fr_coordinate.readline().strip('\n'))
	font_size_17 = int(fr_coordinate.readline().strip('\n'))
	font_size_26 = int(fr_coordinate.readline().strip('\n'))

if not os.path.exists('/home/pi/VE100/.oldemail.txt'):
	fw_email = open('/home/pi/VE100/.oldemail.txt', 'x')
	fw_email.writelines('@gmail.com\n')
	fw_email.close()
	autofill_email = "@gmail.com"
else:
	fr_email = open('/home/pi/VE100/.oldemail.txt')
	autofill_email = fr_email.readline()

if not os.path.exists('/home/pi/VE100/bandcrop.txt'):
	fw_crop = open('/home/pi/VE100/bandcrop.txt', 'x')
	fw_crop.writelines('25\n')
	fw_crop.writelines('203\n')
	fw_crop.writelines('860\n')
	fw_crop.writelines('465\n')
	fw_crop.close()
	band_crop_x1 = 25
	band_crop_y1 = 203
	band_crop_x2 = 860
	band_crop_y2 = 465
else:
	fr_crop = open('/home/pi/VE100/bandcrop.txt')
	band_crop_x1 = int(fr_crop.readline())
	band_crop_y1 = int(fr_crop.readline().strip('\n'))
	band_crop_x2 = int(fr_crop.readline().strip('\n'))
	band_crop_y2 = int(fr_crop.readline().strip('\n'))

# INIT FOLDER - END

##### ACTIVE CODE HANDLE - START #####
try:
	fr = open("/home/pi/VE100/active_code.txt","r")
	active_code = fr.readline().strip('\n')
except:
	active_code = 'phusa@full'
##### ACTIVE CODE HANDLE - END #####
##### TRIAL EXPIRED HANDLE - START #####
try:
	fr = open("/var/tmp/.trial_info.txt","r")
	trial_date = int(fr.readline())
	trial_month = int(fr.readline())
	trial_year = int(fr.readline())
	trial_30days_extend_code = fr.readline().strip('\n')
	trial_full_active_code = fr.readline().strip('\n')
except:
	trial_date = 0
	trial_month = 0
	trial_year = 0
	trial_30days_extend_code = 'phusa@30'
	trial_full_active_code = 'phusa@full'
##### TRIAL EXPIRED HANDLE - END #####

# UART - START
ser = serial.Serial(
	port = '/dev/serial0',
	baudrate = 38400,
	parity = serial.PARITY_NONE,
	stopbits = serial.STOPBITS_ONE,
	bytesize = serial.EIGHTBITS,
	timeout = 1
)

def uart_send(vol, state):
	send_data = '\rVOLTAGE ' + str(vol) + '\r'
	ser.write(send_data.encode())
	sleep(0.1)

	send_data = '\rCURRENT 10\r'
	ser.write(send_data.encode())
	sleep(0.1)

#     send_data = '\rECHO 1'
#     ser.write(send_data.encode())
#     sleep(0.1)

	send_data = '\rCOMMIT\r'
	ser.write(send_data.encode())
	sleep(0.1)

	send_data = '\rOUTPUT ' + str(state) + '\r'
	ser.write(send_data.encode())
	sleep(0.1)
# UART - END

# GPIO - START
GPIO.setwarnings(False)
GPIO.setmode(GPIO.BCM)
GPIO.setup(RELAY_PIN, GPIO.OUT, initial=GPIO.LOW)
GPIO.setup(BLUE_LIGHT_PIN, GPIO.OUT, initial=GPIO.LOW)
GPIO.setup(POWER_LED_PIN, GPIO.OUT, initial=GPIO.HIGH)
GPIO.setup(RUN_LED_PIN, GPIO.OUT, initial=GPIO.LOW) 
GPIO.setup(BUZZER_PIN, GPIO.OUT, initial=GPIO.LOW)
GPIO.setup(SENSOR_PIN, GPIO.IN)
# GPIO - END


# ADS1115 - START
i2c = busio.I2C(3, 2)
ads = ADS.ADS1115(i2c)
# ADS1115 - END

# DS1307
i2c_2 = board.I2C()
rtc = adafruit_ds1307.DS1307(i2c_2)

# CAMERA - START
camera = PiCamera(framerate = 3, sensor_mode = 3)
camera.resolution = (1024,768)
camera.rotation = 0
# camera.brightness = 53
# camera.contrast = 5
def camera_capture(output):
	global camera, final_cap
	# if(final_cap):
	#     camera.close()
	#     sleep(1)
	#     camera = PiCamera(framerate = 3, sensor_mode = 3)
	#     camera.resolution = (1024,768)
	#     sleep(2)
	#     final_cap = 0
	camera.iso = 200
	camera.shutter_speed = 6000000
	camera.exposure_mode = 'off'
	camera.capture(output)

	edit_img = cv2.imread(output)
	alpha = 1.5 # Contrast
	beta = 10 # Brightness
	adjusted = cv2.convertScaleAbs(edit_img, alpha=alpha, beta=beta)
	cv2.imwrite(output, adjusted)

def camera_preview():
	camera.iso = 0
	camera.shutter_speed = 2000000
	camera.exposure_mode = 'off'
	camera.start_preview(alpha=255, fullscreen=False, window=(234, -10, 505, 450))
# CAMERA - END

# EMAIL - START
def sendmail(recipient, subject, content, zip_file):
	global email_password, email_address
	print("email_address: ", email_address)
	print("email_password: ", email_password)

	emailData = MIMEMultipart()
	emailData['Subject'] = subject
	emailData['To'] = recipient
	emailData['From'] = email_address

	emailData.attach(MIMEText(content))

#     imageData = MIMEImage(open(image, 'rb').read(), 'jpg')
#     imageData.add_header('Content-Disposition', 'attachment; filename="image.jpg"')
#     emailData.attach(imageData)

	with open(zip_file,'rb') as file:
		emailData.attach(MIMEApplication(file.read(), Name= folder_name_set + '.zip'))

	session = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
	session.ehlo()
	session.starttls()
	session.ehlo()

	#session.login(mail_address, password)
	session.login(email_address, email_password)

	session.sendmail(email_address, recipient.split(','), emailData.as_string())
	session.quit
# EMAIL - END

# MAIN WINDOW - START
root = Tk()
root.geometry('800x480')
root.configure(background = MAIN_BACKGROUND_COLOR)
root.attributes('-fullscreen', True)
root.resizable(False,False)
# MAIN WINDOW - END

trial_work_frame = Frame(root, bg = 'white')
menu_labelframe =  LabelFrame(root, bg=TASKBAR_BACKGROUND_COLOR, width=799, height=65)
active_code_entry = Entry(trial_work_frame, width=30, font=('Courier',14))

# SCROLL FRAME - START
class ScrollableFrame(Frame):
	def __init__(self, container, *args, **kwargs):
		super().__init__(container, *args, **kwargs)
		canvas = Canvas(self, bg = 'white', height=300, width=470)
		scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
		self.scrollable_frame = Frame(canvas, bg = 'white')

		self.scrollable_frame.bind(
			"<Configure>",
			lambda e: canvas.configure(
				scrollregion=canvas.bbox("all")
			)
		)

		canvas.create_window((0, 0), window=self.scrollable_frame)

		canvas.configure(yscrollcommand=scrollbar.set)

		canvas.pack(side="left", expand=TRUE)
		scrollbar.pack(side="right", fill="y")
# SCROLL FRAME - END

account_labelframe = LabelFrame() 
run_labelframe = LabelFrame()
viewresults_labelframe = LabelFrame()

# MAIN FUNCTION- START
def main():
	line1_value = DoubleVar()
	line2_value = DoubleVar()
	bandline_value = DoubleVar()
	
	global menu_labelframe
	menu_labelframe =  LabelFrame(root, bg=TASKBAR_BACKGROUND_COLOR, width=799, height=65)
	menu_labelframe.place(x=1,y=415)

	def run_clicked():
		run_button['bg'] = MAINMENU_BUTTON_ENABLE_COLOR
		viewresults_button['bg'] = MAINMENU_BUTTON_DISABLE_COLOR
		account_button['bg'] = MAINMENU_BUTTON_DISABLE_COLOR
#         exit_button['bg'] = MAINMENU_BUTTON_DISABLE_COLOR
		
		global account_labelframe, run_labelframe, viewresults_labelframe
		try:
			viewresults_labelframe.destroy()
		except:
			pass
		try:
			account_labelframe.destroy()
		except:
			pass

		run_labelframe =  LabelFrame(root, bg=MAIN_BACKGROUND_COLOR, width=799, height=419)
		run_labelframe.place(x=1,y=1)

		run_label =  Label(run_labelframe, text="RUN", font=("Courier",14, "bold"), fg=LABEL_TEXT_COLOR_1, bg=LABEL_BACKGROUND_COLOR_1, width=72, height=2)
		run_label.place(x=0,y=1)
		ve100_label = Label(run_labelframe, text="VE100", font=("Courier",14, "bold"), fg=LABEL_TEXT_COLOR_2, bg=LABEL_BACKGROUND_COLOR_2, width=10, height=2)
		ve100_label.place(x=0,y=1)

		mode_labelframe = LabelFrame(run_labelframe, bg=MAIN_BACKGROUND_COLOR, width=500, height=200, text="Run Mode")
		mode_labelframe.place(x=146,y=130)

		def singlestage_clicked():
			folder_labelframe = LabelFrame(run_labelframe,
												width = 100,
												height = 50,
												text = "Folder Name",
												bg = 'dodger blue')
			folder_labelframe.place(x=130, y=130)

			folder_name_entry = Entry(folder_labelframe, width=30, justify='right', font=('Courier',14))
			folder_name_entry.pack(padx=100, pady=30)
			folder_name_entry.focus_set()

			def ok_clicked():
				global folder_name_set, preview_only, multi_chosen, single_chosen, path_name, path_name_0
				if(folder_name_entry.get() != ''):
				
					create_time = strftime("%y-%m-%d")
					if not os.path.exists("/home/pi/Desktop/VE100 Result/" + create_time):
						path_name_0 = os.path.join("/home/pi/Desktop/VE100 Result/", create_time)
						os.mkdir(path_name_0)
					else:
						path_name_0 =  "/home/pi/Desktop/VE100 Result/" + create_time
					
					if(os.path.exists(path_name_0 + "/" + folder_name_entry.get())):
						msg = messagebox.askquestion("", "This folder already exists, do you want to overwrite it ?")
						if(msg == 'yes'):
							folder_name_set = folder_name_entry.get()
							path_name = os.path.join(path_name_0, folder_name_set +'/')
							shutil.rmtree(path_name)
							os.mkdir(path_name)
							
							try:
								naming_labelframe.destroy()
							except:
								pass
							single_chosen = 1
							multi_chosen = 0
							preview_only = 0
							menu_labelframe.destroy()
							namingScreen()
					else:
						folder_name_set = folder_name_entry.get()
						path_name = os.path.join(path_name_0, folder_name_set +'/')
						os.mkdir(path_name)
						
						try:
							naming_labelframe.destroy()
						except:
							pass
						single_chosen = 1
						multi_chosen = 0
						preview_only = 0
						menu_labelframe.destroy()
						namingScreen()
						
				else:
					messagebox.showwarning("","Please enter the folder name !")
						
						
			def cancel_clicked():
				folder_labelframe.place_forget()

			ok_button = Button(folder_labelframe,
						text = "OK",
						font = ('Courier', 13),
						width = 5,
						height = 1,
						bg = 'grey85',
						borderwidth = 0,
						command = ok_clicked)
			ok_button.pack(side=LEFT, padx=20, pady=30, ipady=10, ipadx=20)

			cancel_button = Button(folder_labelframe,
						text = "Cancel",
						font = ('Courier', 13),
						width = 5,
						height = 1,
						bg = 'grey85',
						borderwidth = 0,
						command = cancel_clicked)
			cancel_button.pack(side=RIGHT, padx=20, pady=30, ipady=10, ipadx=20)

		def multistage_clicked():
			folder_labelframe = LabelFrame(run_labelframe,
												width = 100,
												height = 50,
												text = "Folder Name",
												bg = 'dodger blue')
			folder_labelframe.place(x=130, y=130)

			folder_name_entry = Entry(folder_labelframe, width=30, justify='right', font=('Courier',14))
			folder_name_entry.pack(padx=100, pady=30)
			folder_name_entry.focus_set()

			def ok_clicked():
				global folder_name_set, multi_chosen, single_chosen, path_name, path_name_0
				
				if(folder_name_entry.get() != ""):
					create_time = strftime("%y-%m-%d")
					if not os.path.exists("/home/pi/Desktop/VE100 Result/" + create_time):
						path_name_0 = os.path.join("/home/pi/Desktop/VE100 Result/", create_time)
						os.mkdir(path_name_0)
					else:
						path_name_0 =  "/home/pi/Desktop/VE100 Result/" + create_time
					
					if(os.path.exists(path_name_0 + "/" + folder_name_entry.get())):
						msg = messagebox.askquestion("", "This folder already exists, do you want to overwrite it ?")
						if(msg == 'yes'):
							folder_name_set = folder_name_entry.get()
							path_name = os.path.join(path_name_0, folder_name_set +'/')
							shutil.rmtree(path_name)
							os.mkdir(path_name)
							
							try:
								naming_labelframe.destroy()
							except:
								pass
							single_chosen = 0
							multi_chosen = 1
							menu_labelframe.destroy()
							namingScreen()
					else:
						folder_name_set = folder_name_entry.get()
						path_name = os.path.join(path_name_0, folder_name_set +'/')
						os.mkdir(path_name)
						
						try:
							naming_labelframe.destroy()
						except:
							pass
						single_chosen = 0
						multi_chosen = 1
						menu_labelframe.destroy()
						namingScreen()
				else:
					messagebox.showwarning("","Please enter Please enter the folder name !")
				

			def cancel_clicked():
				folder_labelframe.place_forget()

			ok_button = Button(folder_labelframe,
						text = "OK",
						font = ('Courier', 13),
						width = 5,
						height = 1,
						bg = 'grey85',
						borderwidth = 0,
						command = ok_clicked)
			ok_button.pack(side=LEFT, padx=20, pady=30, ipady=10, ipadx=20)

			cancel_button = Button(folder_labelframe,
						text = "Cancel",
						font = ('Courier', 13),
						width = 5,
						height = 1,
						bg = 'grey85',
						borderwidth = 0,
						command = cancel_clicked)
			cancel_button.pack(side=RIGHT, padx=20, pady=30, ipady=10, ipadx=20)

		singlestage_button = Button(mode_labelframe, bg=BUTTON_BACKGROUND_COLOR_1, fg=BUTTON_TEXT_COLOR_1, bd=0, text='Single Stage', width=17, height=4, command=singlestage_clicked)
		singlestage_button.place(x=50,y=47)
		multistage_button = Button(mode_labelframe, bg=BUTTON_BACKGROUND_COLOR_1, fg=BUTTON_TEXT_COLOR_1, bd=0, text='Multi Stage', width=17, height=4, command=multistage_clicked)
		multistage_button.place(x=281,y=47)

	def viewresults_clicked():
		run_button['bg'] = MAINMENU_BUTTON_DISABLE_COLOR
		viewresults_button['bg'] = MAINMENU_BUTTON_ENABLE_COLOR
		account_button['bg'] = MAINMENU_BUTTON_DISABLE_COLOR
#         exit_button['bg'] = MAINMENU_BUTTON_DISABLE_COLOR

		global account_labelframe, run_labelframe, viewresults_labelframe
		try:
			run_labelframe.destroy()
		except:
			pass
		try:
			account_labelframe.destroy()
		except:
			pass

		viewresults_labelframe =  LabelFrame(root, bg=MAIN_BACKGROUND_COLOR, width=799, height=419)
		viewresults_labelframe.place(x=1,y=1)
		viewresults_label =  Label(viewresults_labelframe, text="BAND FINDER", font=("Courier",14, "bold"), fg=LABEL_TEXT_COLOR_1, bg=LABEL_BACKGROUND_COLOR_1, width=72, height=2)
		viewresults_label.place(x=0,y=1)
		ve100_label = Label(viewresults_labelframe, text="VE100", font=("Courier",14, "bold"), fg=LABEL_TEXT_COLOR_2, bg=LABEL_BACKGROUND_COLOR_2, width=10, height=2)
		ve100_label.place(x=0,y=1)

		def band_analysis():
			global band_crop_x1, band_crop_x2, band_crop_y1, band_crop_y2, scale_percent_1, scale_percent_2
#             menu_labelframe.place_forget()

			control_labelframe = LabelFrame(viewresults_labelframe, bg=FRAME_BACKGROUND_COLOR_1, width=759, height=70)
			control_labelframe.place(x=18,y=340)
			control1_labelframe = LabelFrame(viewresults_labelframe, bg=FRAME_BACKGROUND_COLOR_1, width=180, height=70)
			control1_labelframe.place(x=18,y=340)
			control2_labelframe = LabelFrame(viewresults_labelframe, bg=FRAME_BACKGROUND_COLOR_1, width=290, height=70)
			control2_labelframe.place(x=198,y=340)
			control3_labelframe = LabelFrame(viewresults_labelframe, bg=FRAME_BACKGROUND_COLOR_1, width=160, height=70)
			control3_labelframe.place(x=488,y=340)
			control4_labelframe = LabelFrame(viewresults_labelframe, text='Note',bg=FRAME_BACKGROUND_COLOR_1, fg='blue', width=130, height=70)
			control4_labelframe.place(x=648,y=340)

			def draw_line_1(value=None):
				global line1, band_line
				try:
					image_canvas.delete(line1)
					image_canvas.delete(band_line)
					image_canvas.delete(band_line)
					band_value_slidebar.place_forget()
				except:
					pass
				line1 = image_canvas.create_line(1, line1_value.get(), 750,  line1_value.get(), fill="dodger blue", width=1)

			def draw_line_2(value=None):
				global line2, band_line
				try:
					image_canvas.delete(line2)
					band_value_slidebar.place_forget()
				except:
					pass
				line2 = image_canvas.create_line(1, line2_value.get(), 750,  line2_value.get(), fill="red", width=1)

			def draw_line_3(value=None):
				global band_line, line1, line2, band_finder_clicked
				try:
					image_canvas.delete(line1)
				except:
					pass
				try:
					image_canvas.delete(line2)
				except:
					pass
				try:
					image_canvas.delete(band_line)
				except:
					pass

				# a_value = (line2_value.get() - line1_value.get())/(math.log(100) - math.log(2000))
#                 b_value = line1_value.get() - math.log(2000)*a_value
				band_distance = bandline_value.get()

				if(band_distance > line2_value.get()):
					band_value_slidebar.set(line2_value.get())
				if(band_distance < line1_value.get()):
					band_value_slidebar.set(line1_value.get())

				tmp_value = (band_distance - b_value)/a_value
				band_value = round(pow(10, tmp_value));
				if(band_value>2000):
					band_value = 2000
				if(band_value<100):
					band_value = 100

				print("a_value: ", a_value)
				print("b_value: ", b_value)
				print("band_distance: ", band_distance)
				print("tmp_value: ", tmp_value)
				print("band_value: ", band_value)

				if(band_finder_clicked == 0):
					band_entry.delete(0,END);
					band_entry.insert(0, band_value)
				else:
					band_finder_clicked = 0

				band_line = image_canvas.create_line(1, bandline_value.get(), 750,  bandline_value.get(), fill="green2", width=1)


			band_value_slidebar = Scale(viewresults_labelframe, variable=bandline_value, from_=1, to=250, bg="green2", orient=VERTICAL, length=280, command=draw_line_3)
			lower_band_slidebar = Scale(viewresults_labelframe, variable=line2_value, from_=1, to=250, bg="red", orient=VERTICAL, length=280, command=draw_line_2)
			lower_band_slidebar.place(x=727, y=56)
			higher_band_slidebar = Scale(viewresults_labelframe, variable=line1_value, from_=1, to=250, bg="dodger blue", orient=VERTICAL, length=280, command=draw_line_1)
			higher_band_slidebar.place(x=707, y=56)

			image0_canvas = Canvas(viewresults_labelframe, bg=FRAME_BACKGROUND_COLOR_1, width=716, height=282)
			image0_canvas.place(x=18, y=56)
			global image_canvas
			image_canvas = Canvas(viewresults_labelframe, bg='grey35', width=699, height=250)
			image_canvas.place(x=35, y=72)

			band_label = Label(control2_labelframe, text="Band size", bg=FRAME_BACKGROUND_COLOR_1)
			band_label.place(x=21,y=4)
			band_entry = Entry(control2_labelframe, state='disable', justify='center',width=7, font=('Courier',20))
			band_entry.place(x=22,y=24)
			band_entry.insert(0, 2000)

			def open_clicked():
				global display_img, higher_band_ypixel, lower_band_ypixel, file_name
				global open_file_name
				open_file_name = filedialog.askopenfilename(initialdir="/home/pi/Desktop/VE100 Result", filetypes=[('png file','*png')])
				if open_file_name is not None:
					if(open_file_name[len(open_file_name)-3:]=='png'):
						img = Image.open(open_file_name)
						crop_width, crop_height = img.size
						scale_percent = scale_percent_1
						width = int(crop_width * scale_percent / 100)
						height = int(crop_height * scale_percent / 100)
						global display_img
						display_img = img.resize((width,height))
						crop_area = (band_crop_x1,band_crop_y1,band_crop_x2,band_crop_y2)
						display_img = display_img.crop(crop_area)
						display_img = ImageTk.PhotoImage(display_img)

						image_canvas.create_image(0, 0, anchor=NW, image=display_img)

						find_button['state'] = 'normal'
						band_entry['state'] = 'normal'
			def find_clicked():
				global a_value, b_value, band_value, band_distance, band_line, band_finder_clicked
				
				band_finder_clicked = 1
				
				if(band_entry.get()==''):
					msg = messagebox.showwarning("", "You haven't chosen a band yet !")
				elif(int(band_entry.get())>2000 or int(band_entry.get())<0):
					msg = messagebox.showwarning("", "Band size must be between 0 and 2000 bp !")
				else:
					save_button['state'] = 'normal'

					a_value = round((line2_value.get() - line1_value.get())/(math.log10(100) - math.log10(2000)),2)
					b_value = round(line1_value.get() - math.log10(2000)*a_value,2)

					band_value = int(band_entry.get())
					#global band_distance
					band_distance = round(math.log10(band_value)*a_value + b_value)
					band_value_slidebar.set(band_distance)

					print("line1_value = ", line1_value.get())
					print("line2_value = ", line2_value.get())
					print("a = ", a_value)
					print("b = ", b_value)
					print("band_value = ", band_value)
					print("band_distance = ", band_distance)

					try:
						image_canvas.delete(line1)
						image_canvas.delete(line2)
						image_canvas.delete(band_line)
					except:
						pass
					band_line = image_canvas.create_line(1,  band_distance , 750, band_distance, fill="lawn green", width=1)

				save_button['state']='normal'
			   #  lower_band_slidebar.place_forget()
#                 higher_band_slidebar.place_forget()
				band_value_slidebar.place(x=746, y=56)

			def save_clicked():
				f = filedialog.asksaveasfile(filetypes=[("eps file", ".jpg")], defaultextension = '.jpg')
				if f is not None:
					name = f.name
					print(name)
					# image_canvas.update()
#                     image_canvas.postscript(file=name,colormode='color')
#                     image_save = Image.open(name)
#                     name = os.path.splitext(name)[0]
#                     image_save.save(name+'.png', 'png')
#                     final_img = Image.open(name +'.png')
#                     crop_width, crop_height = final_img.size
#                     scale_percent = 150
#                     width = int(crop_width * scale_percent / 100)
#                     height = int(crop_height * scale_percent / 100)
#                     final_img = final_img.resize((width,height))
#                     d= ImageDraw.Draw(final_img)
#                     d.text((705,250), band_entry.get() + " bp", fill=(0,255,0))
#                     final_img.save(name+'.png', 'png')

					img = cv2.imread(open_file_name)
					start_point = (0, round((band_distance + band_crop_y1)*scale_percent_2/100))
					end_point = (1024,  round((band_distance + band_crop_y1)*scale_percent_2/100))
					final_img = cv2.line(img, start_point, end_point, (0,255,0), 1)
					final_img = cv2.putText(final_img,band_entry.get()+' bp', (870,730), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,255,0),2, cv2.LINE_AA)
					name = os.path.splitext(name)[0]
					cv2.imwrite(name+'.jpg', final_img)
					msg = messagebox.showinfo("","Saved !")


			open_button = Button(control1_labelframe, text='Open', font=('Courier', 13), fg=BUTTON_TEXT_COLOR_1, bg=BUTTON_BACKGROUND_COLOR_2, borderwidth=2, width=10, height=2, command=open_clicked)
			open_button.place(x=22,y=4)

			find_button = Button(control2_labelframe, state='disable', text='Find', font=('Courier', 13), fg=BUTTON_TEXT_COLOR_1, bg=BUTTON_BACKGROUND_COLOR_2, borderwidth=2, width=8, height=2, command=find_clicked)
			find_button.place(x=151,y=4)
			save_button = Button(control3_labelframe, state='disable',text='Save', font=('Courier', 13), fg=BUTTON_TEXT_COLOR_1, bg=BUTTON_BACKGROUND_COLOR_2, borderwidth=2, width=8, height=2, command=save_clicked)
			save_button.place(x=22,y=4)

			note1_label = Label(control4_labelframe, text="First Band", fg ="dodger blue", bg='grey88')
			note1_label.place(x=27,y=2)
			note2_label = Label(control4_labelframe, text="Last Band", fg ="red", bg='grey88')
			note2_label.place(x=27,y=25)

		band_analysis()

	def account_clicked():
		run_button['bg'] = MAINMENU_BUTTON_DISABLE_COLOR
		viewresults_button['bg'] = MAINMENU_BUTTON_DISABLE_COLOR
		account_button['bg'] = MAINMENU_BUTTON_ENABLE_COLOR
#         exit_button['bg'] = MAINMENU_BUTTON_DISABLE_COLOR

		global account_labelframe, run_labelframe, viewresults_labelframe
		try:
			run_labelframe.destroy()
		except:
			pass
		try:
			viewresults_labelframe.destroy()
		except:
			pass

		account_labelframe =  LabelFrame(root, bg=MAIN_BACKGROUND_COLOR, width=799, height=419)
		account_labelframe.place(x=1,y=1)
		account_label =  Label(account_labelframe, text="ACCOUNT", font=("Courier",14, "bold"), fg=LABEL_TEXT_COLOR_1, bg=LABEL_BACKGROUND_COLOR_1, width=72, height=2)
		account_label.place(x=0,y=1)
		ve100_label = Label(account_labelframe, text="VE100", font=("Courier",14, "bold"), fg=LABEL_TEXT_COLOR_2, bg=LABEL_BACKGROUND_COLOR_2, width=10, height=2)
		ve100_label.place(x=0,y=1)

		email_label = Label(account_labelframe, bg=MAIN_BACKGROUND_COLOR, text='Email', font=('Courier, 15'))
		email_label.place(x=158, y=175)
		email_entry = Entry(account_labelframe, justify='right',width=28, font=('Courier',15))
		email_entry.place(x=268,y=174)
		password_label = Label(account_labelframe, bg=MAIN_BACKGROUND_COLOR, text='Password', font=('Courier, 15'))
		password_label.place(x=158, y=215)
		password_entry = Entry(account_labelframe, show='◼', justify='right',width=28, font=('Courier',15))
		password_entry.place(x=268,y=214)
		login_label = Label(account_labelframe, bg=MAIN_BACKGROUND_COLOR, fg="grey30", font=("Courier", 15))
		login_label['fg'] = "black"

		hide_var = IntVar()
		def hide_charaters():
			if(hide_var.get()==0):
				password_entry['show']=""
#                 text=password_entry.get()
#                 password_entry.delete(0,END)
#                 password_entry.insert(0,text)
			else:
				password_entry['show']="◼"
#                 text=password_entry.get()
#                 password_entry.delete(0,END)
#                 password_entry.insert(0,text)

		hidepass_checkbutton = Checkbutton(account_labelframe, variable=hide_var, bg=MAIN_BACKGROUND_COLOR, text = "Hide charaters",onvalue=1, offvalue=0, command=hide_charaters)
		hidepass_checkbutton.select()
		hidepass_checkbutton.place(x=488,y=245)

		if(account_active):
			login_label['fg'] = "green4"
			login_label['text'] = 'YOU ARE ALREADY LOGGED IN ✔'
			login_label.place(x=240, y=115)
			email_entry.insert(0,email_address)
			password_entry.insert(0,email_password)
			email_entry['state']='disable'
			password_entry['show']="◼"
			password_entry['state']='disable'
			hidepass_checkbutton.place_forget()

			def logout_clicked():
				msg = messagebox.askquestion("LOGOUT","Are you sure you want to logout ?")
				if(msg=='yes'):
					save_file = open("/home/pi/VE100/.account.txt","w")
					save_file.writelines('0' + "\n")
					save_file.writelines("\n")
					save_file.writelines("\n")
					global account_active, email_password, email_address
					account_active = 0
					email_address = '\n'
					email_password = '\n'
					account_clicked()

			logout_button = Button(account_labelframe, bg=BUTTON_BACKGROUND_COLOR_1, bd=0, text='Logout', width=12, height=3, command=logout_clicked)
			logout_button.place(x=333,y=272)
		else:
			login_label['text'] = 'USER LOGIN'
			login_label.place(x=335, y=115)
			def save_click():
				if(email_entry.get()==''):
					messagebox.showwarning("","Please enter the email address!")
				elif(password_entry.get()==''):
					messagebox.showwarning("","Please enter the password !")
				else:
					mail_address = email_entry.get()
					password = password_entry.get()

					addressToVerify = mail_address
					match = re.match('^[_a-z0-9-]+(\.[_a-z0-9-]+)*@[a-z0-9-]+(\.[a-z0-9-]+)*(\.[a-z]{2,4})$', addressToVerify)
					if(match == None):
						messagebox.showerror("","Email syntax error")
					else:
						domain_name = mail_address.split('@')[1]
						records = dns.resolver.query(domain_name, 'MX')
						mxRecord = records[0].exchange
						mxRecord = str(mxRecord)

						host = socket.gethostname()

						server = smtplib.SMTP()
						server.set_debuglevel(0)

						server.connect(mxRecord)
						server.helo(host)
						server.mail('me@domain.com')
						code, message = server.rcpt(str(addressToVerify))
						server.quit()

						if(code==250):
							server=smtplib.SMTP('smtp.gmail.com:587')
							server.starttls()
							try:
								server.login(mail_address,password)
								save_file = open("/home/pi/VE100/.account.txt","w")
								save_file.writelines('1' + "\n")
								save_file.writelines(mail_address + "\n")
								save_file.writelines(password + "\n")
								messagebox.showinfo("", "Saved !")
								global account_active, email_password, email_address
								account_active = 1
								email_address = mail_address
								email_password = password
								account_clicked()
							except:
								messagebox.showerror("","Your password was incorrect\rPlease try again !")
							server.quit()
						else:
							messagebox.showerror("","Your email address was incorrect\rPlease try again !")

			save_button = Button(account_labelframe, bg=BUTTON_BACKGROUND_COLOR_1, bd=0, text='Login', width=12, height=3, command=save_click)
			save_button.place(x=333,y=272)

	def exit_clicked():
#         run_button['bg'] = MAINMENU_BUTTON_DISABLE_COLOR
#         viewresults_button['bg'] = MAINMENU_BUTTON_DISABLE_COLOR
#         account_button['bg'] = MAINMENU_BUTTON_DISABLE_COLOR
#         exit_button['bg'] = MAINMENU_BUTTON_ENABLE_COLOR

#         exit_labelframe =  LabelFrame(root, bg=TASKBAR_COLOR, width=799, height=419)
#         exit_labelframe.place(x=1,y=1)
#         exit_label =  Label(exit_labelframe, text="EXIT", font=("Courier",14, "bold"), fg="white", bg=FUNTION_LABEL_COLOR, width=72, height=2)
#         exit_label.place(x=0,y=1)
#         ve100_label = Label(exit_labelframe , text="VE100", font=("Courier",14, "bold"), fg=DEVIDENAME_TEXT_COLOR, bg=DEVIDENAME_BACKGROUND_COLOR, width=10, height=2)
#         ve100_label.place(x=0,y=1)

		msg = messagebox.askquestion("EXIT", "Do you want to close Application ?")
		if(msg=='yes'):
			os._exit(0)
			root.destroy()
	
	def preview_clicked():
		global preview_only, multi_chosen, single_chosen, menu_labelframe
		single_chosen = 0
		multi_chosen = 0
		preview_only = 1
		menu_labelframe.destroy()
		namingScreen()
		
	run_button = Button(menu_labelframe, text='RUN', font=('Courier', 13), fg='black', bg=MAINMENU_BUTTON_DISABLE_COLOR, borderwidth=0, width=14, height=2, command=run_clicked)
	run_button.place(x=1,y=3)
	viewresults_button = Button(menu_labelframe, text='BAND FINDER', font=('Courier', 13), fg='black', bg=MAINMENU_BUTTON_DISABLE_COLOR, borderwidth=0, width=14, height=2, command=viewresults_clicked)
	viewresults_button.place(x=168,y=3)
	account_button = Button(menu_labelframe, text='ACCOUNT', font=('Courier', 13), fg='black', bg=MAINMENU_BUTTON_DISABLE_COLOR, borderwidth=0, width=14, height=2, command=account_clicked)
	account_button.place(x=335,y=3)
	preview_button = Button(menu_labelframe, text='PREVIEW', font=('Courier', 13), fg='black', bg=MAINMENU_BUTTON_DISABLE_COLOR, borderwidth=0, width=14, height=2, command=preview_clicked)
	preview_button.place(x=502,y=3)
	exit_button = Button(menu_labelframe, text='EXIT', font=('Courier', 13), fg='black', bg='firebrick3', borderwidth=0, width=10, height=2, command=exit_clicked)
	exit_button.place(x=669,y=3)

	run_clicked()
# MAIN FUNCTION - END


n = StringVar()
def namingScreen():
	global naming_labelframe
	naming_labelframe = LabelFrame(root, bg=MAIN_BACKGROUND_COLOR, width=800, height=480)
	naming_labelframe.place(x=0,y=0)
	naming_title_label =  Label(naming_labelframe, text="SETTINGS", font=("Courier",14, "bold"), fg=LABEL_TEXT_COLOR_1, bg=LABEL_BACKGROUND_COLOR_1, width=72, height=2)
	naming_title_label.place(x=0,y=1)

	number_of_wells_label = Label(naming_labelframe, text="Number of wells", font=("Courier",12), bg=MAIN_BACKGROUND_COLOR, fg='black')
	number_of_wells_label.place(x=10, y=55)

	wells_17_labelframe = LabelFrame(naming_labelframe, bg=MAIN_BACKGROUND_COLOR, width=792, height=315)
	wells_frame = ScrollableFrame(wells_17_labelframe)
	wells_frame.pack(pady=5)

	number0_button_list_17 = Button(wells_frame.scrollable_frame,
					fg = 'black',
					font = ("Courier",12),
					text= 'No',
					width=2,
					borderwidth = 0)
	number0_button_list_17.grid(row=0, column=0, sticky=EW, padx=1, pady=1)

	well0_entry_list_17 = Button(wells_frame.scrollable_frame,
					fg = 'black',
					font = ("Courier",12),
					text= 'Well Name',
					width=39,
					borderwidth = 0)
	well0_entry_list_17.grid(row=0, column=1, sticky=EW, padx=1, pady=1)


	well_entry_list_17 = list(range(18))
	number_button_list_17 = list(range(18))
	n=0
	for i in range(0,18):
		number_button_list_17[i] = Button(wells_frame.scrollable_frame,
						fg = 'black',
						font = ("Courier",12),
						text= str(i+1),
						width=2,
						borderwidth = 0)
		number_button_list_17[i].grid(row=i+1, column=0, sticky=EW, padx=1, pady=1)

		well_entry_list_17[i] = Entry(wells_frame.scrollable_frame,
								width=20,
								justify='left',
								font=('Courier',15))
		well_entry_list_17[i].grid(row=i+1, column=1, sticky=EW, padx=1, pady=1)

	well_entry_list_17[0].focus_set()
	well_entry_list_17[0].bind("<Return>",lambda funct:well_entry_list_17[1].focus_set())
	well_entry_list_17[1].bind("<Return>",lambda funct:well_entry_list_17[2].focus_set())
	well_entry_list_17[2].bind("<Return>",lambda funct:well_entry_list_17[3].focus_set())
	well_entry_list_17[3].bind("<Return>",lambda funct:well_entry_list_17[4].focus_set())
	well_entry_list_17[4].bind("<Return>",lambda funct:well_entry_list_17[5].focus_set())
	well_entry_list_17[5].bind("<Return>",lambda funct:well_entry_list_17[6].focus_set())
	well_entry_list_17[6].bind("<Return>",lambda funct:well_entry_list_17[7].focus_set())
	well_entry_list_17[7].bind("<Return>",lambda funct:well_entry_list_17[8].focus_set())
	well_entry_list_17[8].bind("<Return>",lambda funct:well_entry_list_17[9].focus_set())
	well_entry_list_17[9].bind("<Return>",lambda funct:well_entry_list_17[10].focus_set())
	well_entry_list_17[10].bind("<Return>",lambda funct:well_entry_list_17[11].focus_set())
	well_entry_list_17[11].bind("<Return>",lambda funct:well_entry_list_17[12].focus_set())
	well_entry_list_17[12].bind("<Return>",lambda funct:well_entry_list_17[13].focus_set())
	well_entry_list_17[13].bind("<Return>",lambda funct:well_entry_list_17[14].focus_set())
	well_entry_list_17[14].bind("<Return>",lambda funct:well_entry_list_17[15].focus_set())
	well_entry_list_17[15].bind("<Return>",lambda funct:well_entry_list_17[16].focus_set())
	well_entry_list_17[16].bind("<Return>",lambda funct:well_entry_list_17[17].focus_set())
	well_entry_list_17[17].bind("<Return>",lambda funct:well_entry_list_17[0].focus_set())



	wells_26_labelframe = LabelFrame(naming_labelframe, bg=MAIN_BACKGROUND_COLOR, width=792, height=315)
	wells_frame = ScrollableFrame(wells_26_labelframe)
	wells_frame.pack(pady=5)
	number0_button_list_26 = Button(wells_frame.scrollable_frame,
					fg = 'black',
					font = ("Courier",12),
					text= 'No',
					width=2,
					borderwidth = 0)
	number0_button_list_26.grid(row=0, column=0, sticky=EW, padx=1, pady=1)
	well0_entry_list_26 = Button(wells_frame.scrollable_frame,
					fg = 'black',
					font = ("Courier",12),
					text= 'Well Name',
					width=39,
					borderwidth = 0)
	well0_entry_list_26.grid(row=0, column=1, sticky=EW, padx=1, pady=1)
	well_entry_list_26 = list(range(26))
	number_button_list_26 = list(range(26))
	n=0
	for i in range(0,26):
		number_button_list_26[i] = Button(wells_frame.scrollable_frame,
						fg = 'black',
						font = ("Courier",12),
						text= str(i+1),
						width=2,
						borderwidth = 0)
		number_button_list_26[i].grid(row=i+1, column=0, sticky=EW, padx=1, pady=1)

		well_entry_list_26[i] = Entry(wells_frame.scrollable_frame,
								width=20,
								justify='left',
								font=('Courier',15))
		well_entry_list_26[i].grid(row=i+1, column=1, sticky=EW, padx=1, pady=1)

#     def check_entry_1():
#         if(len(well_entry_list_17[i].get().strip()) > 10):
#             messagebox.showwarning("","Well name must be less than 15 characters.")
#         else:
#             well_entry_list_26[1].focus_set()

	well_entry_list_26[0].focus_set()
	well_entry_list_26[0].bind("<Return>",lambda funct:well_entry_list_26[1].focus_set())
	well_entry_list_26[1].bind("<Return>",lambda funct:well_entry_list_26[2].focus_set())
	well_entry_list_26[2].bind("<Return>",lambda funct:well_entry_list_26[3].focus_set())
	well_entry_list_26[3].bind("<Return>",lambda funct:well_entry_list_26[4].focus_set())
	well_entry_list_26[4].bind("<Return>",lambda funct:well_entry_list_26[5].focus_set())
	well_entry_list_26[5].bind("<Return>",lambda funct:well_entry_list_26[6].focus_set())
	well_entry_list_26[6].bind("<Return>",lambda funct:well_entry_list_26[7].focus_set())
	well_entry_list_26[7].bind("<Return>",lambda funct:well_entry_list_26[8].focus_set())
	well_entry_list_26[8].bind("<Return>",lambda funct:well_entry_list_26[9].focus_set())
	well_entry_list_26[9].bind("<Return>",lambda funct:well_entry_list_26[10].focus_set())
	well_entry_list_26[10].bind("<Return>",lambda funct:well_entry_list_26[11].focus_set())
	well_entry_list_26[11].bind("<Return>",lambda funct:well_entry_list_26[12].focus_set())
	well_entry_list_26[12].bind("<Return>",lambda funct:well_entry_list_26[13].focus_set())
	well_entry_list_26[13].bind("<Return>",lambda funct:well_entry_list_26[14].focus_set())
	well_entry_list_26[14].bind("<Return>",lambda funct:well_entry_list_26[15].focus_set())
	well_entry_list_26[15].bind("<Return>",lambda funct:well_entry_list_26[16].focus_set())
	well_entry_list_26[16].bind("<Return>",lambda funct:well_entry_list_26[17].focus_set())
	well_entry_list_26[17].bind("<Return>",lambda funct:well_entry_list_26[18].focus_set())
	well_entry_list_26[18].bind("<Return>",lambda funct:well_entry_list_26[19].focus_set())
	well_entry_list_26[19].bind("<Return>",lambda funct:well_entry_list_26[20].focus_set())
	well_entry_list_26[20].bind("<Return>",lambda funct:well_entry_list_26[21].focus_set())
	well_entry_list_26[21].bind("<Return>",lambda funct:well_entry_list_26[22].focus_set())
	well_entry_list_26[22].bind("<Return>",lambda funct:well_entry_list_26[23].focus_set())
	well_entry_list_26[23].bind("<Return>",lambda funct:well_entry_list_26[24].focus_set())
	well_entry_list_26[24].bind("<Return>",lambda funct:well_entry_list_26[25].focus_set())
	well_entry_list_26[25].bind("<Return>",lambda funct:well_entry_list_26[0].focus_set())
	  #  def ok_clicked_0(event=None):
#         well_entry_list_26[1].focus_set()
#     well_entry_list_26[i].bind("<Return>", ok_clicked_0)

	def check_option(eventObject):
		if(number_of_wells_combobox.current() == 0):
			try:
				wells_26_labelframe.place_forget()
			except:
				pass
			wells_17_labelframe.place(x=152,y=85)

		else:
			try:
				wells_17_labelframe.place_forget()
			except:
				pass
			wells_26_labelframe.place(x=152,y=85)

	number_of_wells_combobox = ttk.Combobox(naming_labelframe, state = "readonly", width = 10, textvariable = n)
	number_of_wells_combobox['values'] = ('18 wells', '26 wells')
	number_of_wells_combobox.place(x=180, y=53)
	number_of_wells_combobox.current(1)
	# number_of_wells_combobox['state'] = 'disable'
	number_of_wells_combobox.bind("<<ComboboxSelected>>", check_option)


	def back_click():
		naming_labelframe.destroy()
		main()
	def next_click():
		global well_name_list_17, well_name_list_26, num_well
		err = 0
		if(number_of_wells_combobox.current() == 0):
			num_well = 17
			well_name_list_17 = list(range(18))
			for i in range(0,18):
				if(len(well_entry_list_17[i].get().strip()) > 12):
					messagebox.showwarning("","Well name must be less than 12 characters.\n[Well " + str(i+1) + ']')
					err = 1
					break
				else:
					well_name_list_17[i] = well_entry_list_17[i].get().strip()
		else:
			num_well = 26
			well_name_list_26 = list(range(26))
			for i in range(0,26):
				if(len(well_entry_list_26[i].get().strip()) > 12):
					messagebox.showwarning("","Well name must be less than 12 characters.\n[Well " + str(i+1) + ']')
					err = 1
					break
				else:
					well_name_list_26[i] = well_entry_list_26[i].get().strip()
		if(err==0):
			if(preview_only == 1):
				previewScreen()
			else:
				if(multi_chosen == 1):
					multiStepRunScreen()
				else:
					oneStepRunScreen()
	back_button = Button(naming_labelframe, bd=1, bg=BUTTON_BACKGROUND_COLOR_1, font=('Courier',12, 'bold'), text='BACK', width=8, height=3, command=back_click)
	back_button.place(x=0,y=405)
	next_button = Button(naming_labelframe, bg=BUTTON_BACKGROUND_COLOR_1, font=('Courier',12, 'bold'), bd=1,text='NEXT', width=8, height=3, command=next_click)
	next_button.place(x=687,y=405)

	check_option(1)

# RUN FUNCTION - START
def multiStepRunScreen():
	global automail2_is_on

	fr_para2 = open("/home/pi/VE100/parameters2.txt","r")
	voltage1_set = int(fr_para2.readline())
	timer1_set = int(fr_para2.readline().strip('\n'))
	voltage2_set = int(fr_para2.readline().strip('\n'))
	timer2_set = int(fr_para2.readline().strip('\n'))
	voltage3_set = int(fr_para2.readline().strip('\n'))
	timer3_set = int(fr_para2.readline().strip('\n'))
	automail2_is_on = int(fr_para2.readline().strip('\n'))

	m1_set = timer1_set//100
	s1_set = timer1_set - (m1_set*100)
	m2_set = timer2_set//100
	s2_set = timer2_set - (m2_set*100)
	m3_set = timer3_set//100
	s3_set = timer3_set - (m3_set*100)

	multirunscreen_labelframe = LabelFrame(root, bg=MAIN_BACKGROUND_COLOR, width=800, height=480)
	multirunscreen_labelframe.place(x=0,y=0)
	multirunscreen_title_label =  Label(multirunscreen_labelframe, text="SETTINGS", font=("Courier",14, "bold"), fg=LABEL_TEXT_COLOR_1, bg=LABEL_BACKGROUND_COLOR_1, width=72, height=2)
	multirunscreen_title_label.place(x=0,y=1)


	setting1_labelframe = LabelFrame(multirunscreen_labelframe, bg=FRAME_BACKGROUND_COLOR_1, fg=FRAME_TEXT_COLOR_1, text='STAGE 1', font=('Courier',15), width=263, height=355)
	setting1_labelframe.place(x=0,y=51)
	setting2_labelframe = LabelFrame(multirunscreen_labelframe, bg=FRAME_BACKGROUND_COLOR_1, fg=FRAME_TEXT_COLOR_1, text='STAGE 2', font=('Courier',15), width=263, height=355)
	setting2_labelframe.place(x=266,y=51)
	setting3_labelframe = LabelFrame(multirunscreen_labelframe, bg=FRAME_BACKGROUND_COLOR_1, fg=FRAME_TEXT_COLOR_1, text='STAGE 3', font=('Courier',15), width=264, height=355)
	setting3_labelframe.place(x=532,y=51)

	voltage1_labelframe = LabelFrame(setting1_labelframe, bg=FRAME_BACKGROUND_COLOR_2, fg=FRAME_TEXT_COLOR_2, text='VOLTAGE (VDC)', font=('Courier',12), width=238, height=100)
	voltage1_labelframe.place(x=10,y=50)
	timer1_labelframe = LabelFrame(setting1_labelframe, bg=FRAME_BACKGROUND_COLOR_2, fg=FRAME_TEXT_COLOR_2, text='RUN TIMER (min:sec)', font=('Courier',12), width=238, height=100)
	timer1_labelframe.place(x=10,y=168)
	twodot_label = Label(timer1_labelframe, fg=LABEL_TEXT_COLOR_3, bg=LABEL_BACKGROUND_COLOR_3, text =':', font=('Courier', 20))
	twodot_label.place(x=106, y=15)
#     autocap1_labelframe = LabelFrame(setting1_labelframe, bg='grey90', fg='black', text='CAPTURE TIMER (min)', font=('Courier',12), width=238, height=100)
#     autocap1_labelframe.place(x=10,y=218)

	voltage2_labelframe = LabelFrame(setting2_labelframe, bg=FRAME_BACKGROUND_COLOR_2, fg=FRAME_TEXT_COLOR_2, text='VOLTAGE (VDC)', font=('Courier',12), width=238, height=100)
	voltage2_labelframe.place(x=10,y=50)
	timer2_labelframe = LabelFrame(setting2_labelframe, bg=FRAME_BACKGROUND_COLOR_2, fg=FRAME_TEXT_COLOR_2, text='RUN TIMER (min:sec)', font=('Courier',12), width=238, height=100)
	timer2_labelframe.place(x=10,y=168)
	twodot_label = Label(timer2_labelframe, fg=LABEL_TEXT_COLOR_3, bg=LABEL_BACKGROUND_COLOR_3, text =':', font=('Courier', 20))
	twodot_label.place(x=106, y=15)
#     autocap2_labelframe = LabelFrame(setting2_labelframe, bg='grey90', fg='black', text='CAPTURE TIMER (min)', font=('Courier',12), width=238, height=100)
#     autocap2_labelframe.place(x=10,y=218)

	voltage3_labelframe = LabelFrame(setting3_labelframe, bg=FRAME_BACKGROUND_COLOR_2, fg=FRAME_TEXT_COLOR_2, text='VOLTAGE (VDC)', font=('Courier',12), width=238, height=100)
	voltage3_labelframe.place(x=10,y=50)
	timer3_labelframe = LabelFrame(setting3_labelframe, bg=FRAME_BACKGROUND_COLOR_2, fg=FRAME_TEXT_COLOR_2, text='RUN TIMER (min:sec)', font=('Courier',12), width=238, height=100)
	timer3_labelframe.place(x=10,y=168)
	twodot_label = Label(timer3_labelframe, fg=LABEL_TEXT_COLOR_3, bg=FRAME_BACKGROUND_COLOR_2, text =':', font=('Courier', 20))
	twodot_label.place(x=106, y=15)
#     autocap3_labelframe = LabelFrame(setting3_labelframe, bg='grey90', fg='black', text='CAPTURE TIMER (min)', font=('Courier',12), width=238, height=100)
#     autocap3_labelframe.place(x=10,y=218)

	voltage1_entry = Entry(voltage1_labelframe, fg='red', justify='center',width=3, font=('Courier',25))
	voltage1_entry.place(x=84,y=12)
	voltage1_entry.insert(0, voltage1_set)
	m1_entry = Entry(timer1_labelframe, fg='red', justify='center',width=3, font=('Courier',25))
	m1_entry.place(x=31,y=12)
	m1_entry.insert(0, str('%02d'%m1_set))
	s1_entry = Entry(timer1_labelframe, fg='red', justify='center',width=3, font=('Courier',25))
	s1_entry.place(x=135,y=12)
	s1_entry.insert(0, str('%02d'%s1_set))
#     autocap1_entry = Entry(autocap1_labelframe, fg='red', justify='center',width=3, font=('Courier',25))
#     autocap1_entry.place(x=84,y=12)
#     autocap1_entry.insert(0, 5)

	voltage2_entry = Entry(voltage2_labelframe, fg='red', justify='center',width=3, font=('Courier',25))
	voltage2_entry.place(x=84,y=12)
	voltage2_entry.insert(0, voltage2_set)
	m2_entry = Entry(timer2_labelframe, fg='red', justify='center',width=3, font=('Courier',25))
	m2_entry.place(x=31,y=12)
	m2_entry.insert(0, str('%02d'%m2_set))
	s2_entry = Entry(timer2_labelframe, fg='red', justify='center',width=3, font=('Courier',25))
	s2_entry.place(x=135,y=12)
	s2_entry.insert(0, str('%02d'%s2_set))
#     autocap2_entry = Entry(autocap2_labelframe, fg='red', justify='center',width=3, font=('Courier',25))
#     autocap2_entry.place(x=84,y=12)
#     autocap2_entry.insert(0, 5)

	voltage3_entry = Entry(voltage3_labelframe, fg='red', justify='center',width=3, font=('Courier',25))
	voltage3_entry.place(x=84,y=12)
	voltage3_entry.insert(0, voltage3_set)
	m3_entry = Entry(timer3_labelframe, fg='red', justify='center',width=3, font=('Courier',25))
	m3_entry.place(x=31,y=12)
	m3_entry.insert(0, str('%02d'%m3_set))
	s3_entry = Entry(timer3_labelframe, fg='red', justify='center',width=3, font=('Courier',25))
	s3_entry.place(x=135,y=12)
	s3_entry.insert(0, str('%02d'%s3_set))
#     autocap3_entry = Entry(autocap3_labelframe, fg='red', justify='center',width=3, font=('Courier',25))
#     autocap3_entry.place(x=84,y=12)
#     autocap3_entry.insert(0, 5)

	automail_labelframe = LabelFrame(multirunscreen_labelframe, bg=FRAME_BACKGROUND_COLOR_2, text='AUTO SEND EMAIL', width=467, height=68)
	automail_labelframe.place(x=219,y=407)

	mail_label = Label(automail_labelframe, bg=FRAME_BACKGROUND_COLOR_2, text='To:', font=('Courier, 15'))
	mail_label.place(x=130, y=7)
	mail_entry = Entry(automail_labelframe, justify='right',width=28, font=('Courier',13))
	mail_entry.place(x=171,y=7)
	mail_entry.insert(0, autofill_email)

	def automail_on_click():
		global automail2_is_on
		automail2_is_on = 1
		automail_on_button['bg'] = 'lawn green'
		automail_off_button['bg'] = MAINMENU_BUTTON_DISABLE_COLOR
		mail_entry['state'] = 'normal'
		if(mail_entry.get()==''):
			mail_entry.insert(0,autofill_email)
		mail_label['fg']='black'
	def automail_off_click():
		global automail2_is_on
		automail2_is_on = 0
		automail_on_button['bg'] = MAINMENU_BUTTON_DISABLE_COLOR
		automail_off_button['bg'] = 'red'
		mail_entry.delete(0,END)
		mail_entry['state'] = 'disable'
		mail_label['fg']='grey75'

	automail_on_button = Button(automail_labelframe,bd=0,text='ON', width=4, height=2, command=automail_on_click)
	automail_on_button.place(x=5,y=-1)
	automail_off_button = Button(automail_labelframe, bg=FRAME_BACKGROUND_COLOR_2, bd=0, text='OFF', width=4, height=2, command=automail_off_click)
	automail_off_button.place(x=63,y=-1)
	if(account_active==0):
		automail_on_button['state'] = 'disable'
		automail_off_button['state'] = 'disable'

	if(automail2_is_on):
		automail_on_click()
	else:
		automail_off_click()

	def run1_click():
		global voltage1_set, voltage2_set, voltage3_set, m1_set, s1_set, m2_set, s2_set, m3_set, s3_set, m0_raw, s0_raw, m1_raw, s1_raw, m2_raw, s2_raw, m3_raw, s3_raw, recipient_email, folder_name, path_name
		global current_measured, folder_name_set
		current_measured = 0

		if(voltage1_entry.get()=='' or voltage2_entry.get()=='' or voltage3_entry.get()==''):
			messagebox.showwarning("","Please enter the voltage !")
		elif((voltage1_entry.get().isnumeric())==0 or (voltage2_entry.get().isnumeric())==0 or (voltage3_entry.get().isnumeric())==0):
			messagebox.showwarning("","Voltage must be between 12 and 80 VDC")
		elif(int(voltage1_entry.get())<12):
			messagebox.showwarning("","Voltage must be between 12 and 80 VDC")
			voltage1_entry.delete(0,END)
			voltage1_entry.insert(0, 12)
		elif(int(voltage1_entry.get())>80):
			messagebox.showwarning("","Voltage must be between 12 and 80 VDC")
			voltage1_entry.delete(0,END)
			voltage1_entry.insert(0, 80)
		elif(int(voltage2_entry.get())<12):
			messagebox.showwarning("","Voltage must be between 12 and 80 VDC")
			voltage2_entry.delete(0,END)
			voltage2_entry.insert(0, 12)
		elif(int(voltage2_entry.get())>80):
			messagebox.showwarning("","Voltage must be between 12 and 80 VDC")
			voltage2_entry.delete(0,END)
			voltage2_entry.insert(0, 80)
		elif(int(voltage3_entry.get())<12):
			messagebox.showwarning("","Voltage must be between 12 and 80 VDC")
			voltage3_entry.delete(0,END)
			voltage3_entry.insert(0, 12)
		elif(int(voltage3_entry.get())>80):
			messagebox.showwarning("","Voltage must be between 12 and 80 VDC")
			voltage3_entry.delete(0,END)
			voltage3_entry.insert(0, 80)
		elif(m1_entry.get()=='' or m2_entry.get()=='' or m3_entry.get()=='' or
				s1_entry.get()=='' or s2_entry.get()=='' or s3_entry.get()==''):
			messagebox.showwarning("","Please enter the timer !")
		elif((mail_entry.get()=='' or mail_entry.get()=='@gmail.com') and automail2_is_on):
			messagebox.showwarning("","Please enter the recipient email !")
		else:
			global running
			running = 1

			global autofill_email
			if(automail2_is_on):
				recipient_email = mail_entry.get()
				fw_email = open('/home/pi/VE100/.oldemail.txt', 'w')
				fw_email.writelines(recipient_email + '\n')
				fw_email.close()
				autofill_email = recipient_email

			voltage1_set = voltage1_entry.get()
			voltage2_set = voltage2_entry.get()
			voltage3_set = voltage3_entry.get()
			m1_set = int(m1_entry.get())
			m2_set = int(m2_entry.get())
			m3_set = int(m3_entry.get())
			s1_set = int(s1_entry.get())
			s2_set = int(s2_entry.get())
			s3_set = int(s3_entry.get())
			m0_raw = m0_set
			s0_raw = s0_set
			m1_raw = m1_set
			s1_raw = s1_set
			m2_raw = m2_set
			s2_raw = s2_set
			m3_raw = m3_set
			s3_raw = s3_set

			subprocess.call(["scrot",path_name + 'parameters.jpg'])

			multirunscreen_labelframe = LabelFrame(root, bg=FRAME_BACKGROUND_COLOR_2, width=800, height=480)
			multirunscreen_labelframe.place(x=0,y=0)
			multirunscreen0_title_labelframe = LabelFrame(multirunscreen_labelframe, text="DEFAULT STAGE", font=('Courier', 13, 'bold'), bg='green yellow', width=170, height=118)
			multirunscreen0_title_labelframe.place(x=0,y=1)
			multirunscreen1_title_labelframe = LabelFrame(multirunscreen_labelframe, text="STAGE 1", font=('Courier', 13, 'bold'), bg='grey70', width=170, height=118)
			multirunscreen1_title_labelframe.place(x=0,y=120)
			multirunscreen2_title_labelframe = LabelFrame(multirunscreen_labelframe, text="STAGE 2", font=('Courier', 13, 'bold'), bg='grey70', width=170, height=118)
			multirunscreen2_title_labelframe.place(x=0,y=239)
			multirunscreen3_title_labelframe = LabelFrame(multirunscreen_labelframe, text="STAGE 3", font=('Courier', 13, 'bold'), bg='grey70', width=170, height=118)
			multirunscreen3_title_labelframe.place(x=0,y=358)

			voltage0_labelframe = LabelFrame(multirunscreen0_title_labelframe, font=('Courier', 12), bg='green yellow', width=140, height=45)
			voltage0_labelframe.place(x=13,y=-5)
			timer0_labelframe = LabelFrame(multirunscreen0_title_labelframe, text="Time Left", font=('Courier', 12), bg='green yellow', width=140, height=45)
			timer0_labelframe.place(x=13,y=42)
			voltage0_label = Label(voltage0_labelframe, fg ='red2', text=str(voltage0_set) + 'V / ' + str('%02d'%m0_raw) + ':' + str('%02d'%s0_raw), font=('calibri',14,'bold'), bg='green yellow')
			voltage0_label.place(x=22, y=11)
			m0_label = Label(timer0_labelframe, fg ='red2', text=str('%.02d'%m0_set), font=('calibri',14,'bold'), bg='green yellow')
			m0_label.place(x=35, y=-2)
			twodot0_label = Label(timer0_labelframe, fg ='red2', text=':', font=('calibri',14,'bold'), bg='green yellow')
			twodot0_label.place(x=63,y=-2)
			s0_label = Label(timer0_labelframe, fg ='red2', text=str('%.02d'%s0_set), font=('calibri',14,'bold'), bg='green yellow')
			s0_label.place(x=77, y=-2)

			voltage1_labelframe = LabelFrame(multirunscreen1_title_labelframe, font=('Courier', 12), bg='grey70', width=140, height=45)
			voltage1_labelframe.place(x=13,y=-5)
			timer1_labelframe = LabelFrame(multirunscreen1_title_labelframe, text="Time Left", font=('Courier', 12), bg='grey70', width=140, height=45)
			timer1_labelframe.place(x=13,y=42)
			voltage1_label = Label(voltage1_labelframe, fg ='red2', text=str(voltage1_set) + 'V / ' + str('%02d'%m1_raw) + ':' + str('%02d'%s1_raw), font=('calibri',14,'bold'), bg='grey70')
			voltage1_label.place(x=22, y=11)
			m1_label = Label(timer1_labelframe, fg ='red2', text=str('%.02d'%m1_set), font=('calibri',14,'bold'), bg='grey70')
			m1_label.place(x=35, y=-2)
			twodot1_label = Label(timer1_labelframe, fg ='red2', text=':', font=('calibri',14,'bold'), bg='grey70')
			twodot1_label.place(x=63,y=-2)
			s1_label = Label(timer1_labelframe, fg ='red2', text=str('%.02d'%s1_set), font=('calibri',14,'bold'), bg='grey70')
			s1_label.place(x=77, y=-2)

			voltage2_labelframe = LabelFrame(multirunscreen2_title_labelframe, font=('Courier', 12), bg='grey70', width=140, height=45)
			voltage2_labelframe.place(x=13,y=-5)
			timer2_labelframe = LabelFrame(multirunscreen2_title_labelframe, text="Time Left", font=('Courier', 12), bg='grey70', width=140, height=45)
			timer2_labelframe.place(x=13,y=42)
			voltage2_label = Label(voltage2_labelframe, fg ='red2', text=str(voltage2_set) + 'V / ' + str('%02d'%m2_raw) + ':' + str('%02d'%s2_raw), font=('calibri',14,'bold'), bg='grey70')
			voltage2_label.place(x=22, y=11)
			m2_label = Label(timer2_labelframe, fg ='red2', text=str('%.02d'%m2_set), font=('calibri',14,'bold'), bg='grey70')
			m2_label.place(x=35, y=-2)
			twodot2_label = Label(timer2_labelframe, fg ='red2', text=':', font=('calibri',14,'bold'), bg='grey70')
			twodot2_label.place(x=63,y=-2)
			s2_label = Label(timer2_labelframe, fg ='red2', text=str('%.02d'%s2_set), font=('calibri',14,'bold'), bg='grey70')
			s2_label.place(x=77, y=-2)

			voltage3_labelframe = LabelFrame(multirunscreen3_title_labelframe, font=('Courier', 12), bg='grey70', width=140, height=45)
			voltage3_labelframe.place(x=13,y=-5)
			timer3_labelframe = LabelFrame(multirunscreen3_title_labelframe, text="Time Left", font=('Courier', 12), bg='grey70', width=140, height=45)
			timer3_labelframe.place(x=13,y=42)
			voltage3_label = Label(voltage3_labelframe, fg ='red2', text=str(voltage3_set) + 'V / ' + str('%02d'%m3_raw) + ':' + str('%02d'%s3_raw), font=('calibri',14,'bold'), bg='grey70')
			voltage3_label.place(x=22, y=11)
			m3_label = Label(timer3_labelframe, fg ='red2', text=str('%.02d'%m3_set), font=('calibri',14,'bold'), bg='grey70')
			m3_label.place(x=35, y=-2)
			twodot3_label = Label(timer3_labelframe, fg ='red2', text=':', font=('calibri',14,'bold'), bg='grey70')
			twodot3_label.place(x=63,y=-2)
			s3_label = Label(timer3_labelframe, fg ='red2', text=str('%.02d'%s3_set), font=('calibri',14,'bold'), bg='grey70')
			s3_label.place(x=77, y=-2)

			preview_labelframe = LabelFrame(multirunscreen_labelframe, bg='black', fg='red', text="◉ MONITOR", font=("Courier",13,'bold'), width=622, height=410)
			preview_labelframe.place(x=173,y=0)

			limit1_canvas = Canvas(preview_labelframe, bg='grey70', bd=0, width=54, height=376)
			limit1_canvas.place(x=0,y=1)
			limit2_canvas = Canvas(preview_labelframe, bg='grey70', bd=0, width=54, height=376)
			limit2_canvas.place(x=562,y=1)
			limit_line1 = limit1_canvas.create_line(x0_limit, y0_limit, x1_limit, y1_limit, fill='red', width=2)
			limit_line2 = limit2_canvas.create_line(x0_limit, y0_limit, x1_limit, y1_limit, fill='red', width=2)

			t_progressbar = atk.RadialProgressbar(preview_labelframe, fg='cyan')
			t_progressbar.place(x=260,y=130)
			t_progressbar.start()
			tprocess_label = Label(preview_labelframe, bg='black', fg='white smoke', text='Processing\r...', font=("Courier",9,'bold'))
			tprocess_label.place(x=272,y=172)

			def time_s3():
				multirunscreen2_title_labelframe['bg']='grey70'
				voltage2_labelframe['bg'] = 'grey70'
				voltage2_label['bg']='grey70'
				timer2_labelframe['bg'] = 'grey70'
				m2_label['bg'] = 'grey70'
				twodot2_label['bg'] = 'grey70'
				s2_label['bg'] = 'grey70'

				multirunscreen3_title_labelframe['bg']='green yellow'
				voltage3_labelframe['bg'] = 'green yellow'
				voltage3_label['bg']='green yellow'
				timer3_labelframe['bg'] = 'green yellow'
				m3_label['bg'] = 'green yellow'
				twodot3_label['bg'] = 'green yellow'
				s3_label['bg'] = 'green yellow'

				global final_cap, s3_set, m3_set, s0_set, m0_set, stage0_is_running, stage3_is_running, result_path, path_name, camera, solve_s3
				global current_measured
				global x_coordinate_17, x_coordinate_26, y_coordinate

		
				s3_set = s3_set - 1
				if(s3_set<0):
					m3_set=m3_set-1
					s3_set=59
				s3_label.config(text = str('%02d'%s3_set))
				m3_label.config(text = str('%02d'%m3_set))

				adc = AnalogIn(ads, ADS.P2)
				adc_voltage = adc.voltage
				real_voltage = round(adc_voltage*max_voltage/divide_voltage)
				present_voltage_label.config(text="Voltage: " + str(real_voltage) + ' V')

				send_data = '\rSTATUS\r'
				ser.write(send_data.encode())
				if(ser.in_waiting>0):
					try:
						current_measured = float(ser.readline().decode('utf-8').rstrip())
					except:
						current_measured = 0.045
						pass
					print("Data received:", current_measured)
					if(current_measured < CURRENT_WARNING):
						present_current_label.config(text="LOW CURRENT WARNING !")
					else:
						present_current_label.config(text="")

				if(m3_set!=-1):
					solve_s3 = s3_label.after(1000, time_s3)
				else:
					limit1_canvas.place_forget()
					limit2_canvas.place_forget()
					present_voltage_label.config(text="Voltage: 0 V")
					present_current_label.config(text="")

					GPIO.output(RELAY_PIN, GPIO.LOW)
					GPIO.output(POWER_LED_PIN, GPIO.LOW)
					uart_send(0,0)

					m3_label.config(text = '00')
					s3_label.config(text = '00')

					try:
						s3_label.after_cancel(solve_s3)
					except:
						pass

#                         camera.close()
#                         sleep(1)
					final_cap = 1
					try:
						camera_capture(path_name + 'stage3_result.png')

						# ~ fr_coordinate = open('/home/pi/VE100/coordinates.txt')
						# ~ x_coordinate_17 = int(fr_coordinate.readline())
						# ~ x_coordinate_26 = int(fr_coordinate.readline().strip('\n'))
						# ~ y_coordinate = int(fr_coordinate.readline().strip('\n'))

						edit_img = Image.open(path_name + 'stage3_result.png')
						img = ImageDraw.Draw(edit_img)
						shape = [(1024, 768), (0,550)]
						img.rectangle(shape, fill ="lightgray", outline="lightgray")
						img_font_1_17 = ImageFont.truetype("/home/pi/VE100/arial.ttf", font_size_17)
						img_font_1_26 = ImageFont.truetype("/home/pi/VE100/arial.ttf", font_size_26)
						img_font_2 = ImageFont.truetype("/home/pi/VE100/arial.ttf", 23)
						if(num_well==17):
							x_coordinate = x_coordinate_17
							for i in range(0,9):
								img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_17, fill=(0,255,0))
								x_coordinate += well_distance_17

							x_coordinate = x_coordinate - well_distance_17 + pace_17
							for i in range(9,18):
								img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_17, fill=(0,255,0))
								x_coordinate += well_distance_17
						else:
							x_coordinate = x_coordinate_26
							for i in range(0,13):
								img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_26, fill=(0,255,0))
								x_coordinate += well_distance_26

							x_coordinate = x_coordinate - well_distance_26 + pace_26
							for i in range(13,26):
								img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_26, fill=(0,255,0))
								# ~ if(i<18):
									# ~ x_coordinate += 26
								# ~ elif(i==18):
									# ~ x_coordinate += 30
								# ~ else:
								x_coordinate += well_distance_26

						if(num_well==26):
							r1 = 568
							r2 = 568
							r3 = 568
							r4 = 568
							r5 = 568
							for i in range(0,26):
								if(i<6):
									img.text((32,r1), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
									r1 += 32
								elif(i<12):
									img.text((232,r2), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
									r2 += 32
								elif(i<18):
									img.text((432,r3), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
									r3 += 32
								elif(i<24):
									img.text((632,r4), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
									r4 += 32
								else:
									img.text((832,r5), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
									r5 += 32
						else:
							r1 = 568
							r2 = 568
							r3 = 568
							r4 = 568
							r5 = 568
							for i in range(0,18):
								if(i<6):
									img.text((32,r1), str(i+1) + '. ' + str(well_name_list_17[i]), font=img_font_2, fill=(0,0,0))
									r1 += 32
								elif(i<12):
									img.text((232,r2), str(i+1) + '. ' + str(well_name_list_17[i]), font=img_font_2, fill=(0,0,0))
									r2 += 32
								else:
									img.text((432,r3), str(i+1) + '. ' + str(well_name_list_17[i]), font=img_font_2, fill=(0,0,0))
									r3 += 32
						edit_img.save(path_name + 'edit_img.png','png')

						# wb = Workbook()
#                         sheet = wb.active
#                         img = Img(path_name + 'edit_img.png')
#                         img.height = 384
#                         img.width = 512
#                         img.anchor = 'B2'
#                         sheet.add_image(img)

#                         sheet.column_dimensions['J'].width = 18
#                         sheet.column_dimensions['K'].width = 18

#                         if(num_well==17):
#                             pass
#                         else:
#                             for i in range(0,13):
#                                 sheet['J'+ str(i+2)] = str(i+1) + '. ' + str(well_name_list_26[i])
#                                 sheet['K'+ str(i+2)] = str(i+14) + '. ' + str(well_name_list_26[i+13])

#                         wb.save(path_name + 'result.xlsx')

					except Exception as e:
						error = messagebox.showerror("ERR 03", str(e), icon = "error")
						if(error=='ok'):
							pass

					m0_set = m0_raw
					s0_set = s0_raw

					multirunscreen3_title_labelframe['bg']='grey70'
					voltage3_labelframe['bg'] = 'grey70'
					voltage3_label['bg']='grey70'
					timer3_labelframe['bg'] = 'grey70'
					m3_label['bg'] = 'grey70'
					twodot3_label['bg'] = 'grey70'
					s3_label['bg'] = 'grey70'

					print("account_active:", account_active)
					print("automail2_is_on:", automail2_is_on)
					if(account_active and automail2_is_on):
						shutil.make_archive(path_name, format='zip', root_dir = path_name)
						try:
							sendmail(recipient_email, folder_name_set, 'This is an email from EV100 device.', path_name_0 + '/' + folder_name_set + '.zip')
						except Exception as e:
							try:
								camera.stop_preview()
							except:
								pass
							error = messagebox.showerror("ERR 02", str(e), icon = "error")
							if(error=='ok'):
								pass
					t_progressbar.place_forget()
					tprocess_label['font'] = ('Courier',20, 'bold')
					tprocess_label['text'] = 'COMPLETED'
					tprocess_label['fg'] = 'lawn green'
					tprocess_label.place(x=236,y=160)

					capture_button['text'] = 'VIEW RESULT'
					stop_button['text'] = 'FINISH'

					stage3_is_running = 0
					stage0_is_running = 1
					try:
						camera.stop_preview()
					except:
						pass
					GPIO.output(BLUE_LIGHT_PIN, GPIO.LOW)
					GPIO.output(RELAY_PIN, GPIO.LOW)
					GPIO.output(RUN_LED_PIN, GPIO.LOW)
					for i in range(0,3):
						GPIO.output(BUZZER_PIN, GPIO.HIGH)
						sleep(0.5)
						GPIO.output(BUZZER_PIN, GPIO.LOW)
						sleep(0.5)
					global running
					running = 0

			def time_s2():
				multirunscreen1_title_labelframe['bg']='grey70'
				voltage1_labelframe['bg'] = 'grey70'
				voltage1_label['bg']='grey70'
				timer1_labelframe['bg'] = 'grey70'
				m1_label['bg'] = 'grey70'
				twodot1_label['bg'] = 'grey70'
				s1_label['bg'] = 'grey70'

				multirunscreen2_title_labelframe['bg']='green yellow'
				voltage2_labelframe['bg'] = 'green yellow'
				voltage2_label['bg']='green yellow'
				timer2_labelframe['bg'] = 'green yellow'
				m2_label['bg'] = 'green yellow'
				twodot2_label['bg'] = 'green yellow'
				s2_label['bg'] = 'green yellow'

				global s2_set, m2_set, stage2_is_running, stage3_is_running, result_path, path_name, camera, solve_s2
				global current_measured

				s2_set = s2_set - 1
				if(s2_set<0):
					m2_set=m2_set-1
					s2_set=59
				s2_label.config(text = str('%02d'%s2_set))
				m2_label.config(text = str('%02d'%m2_set))

				adc = AnalogIn(ads, ADS.P2)
				adc_voltage = adc.voltage
				real_voltage = round(adc_voltage*max_voltage/divide_voltage)
				present_voltage_label.config(text="Voltage: " + str(real_voltage) + ' V')

				send_data = '\rSTATUS\r'
				ser.write(send_data.encode())
				if(ser.in_waiting>0):
					try:
						current_measured = float(ser.readline().decode('utf-8').rstrip())
					except:
						current_measured = 0.045
						pass
					print("Data received:", current_measured)
					if(current_measured < CURRENT_WARNING):
						present_current_label.config(text="LOW CURRENT WARNING !")
					else:
						present_current_label.config(text="")

				if(m2_set!=-1):
					solve_s2 = s2_label.after(1000, time_s2)
				else:
					m2_label.config(text = '00')
					s2_label.config(text = '00')
					try:
						s2_label.after_cancel(solve_s2)
					except:
						pass

					if(s2_raw!=0 or m2_raw!=0):
						try:
							camera_capture(path_name + 'stage2_result.png')
						except Exception as e:
							error = messagebox.showerror("ERR 03", str(e), icon = "error")
							if(error=='ok'):
								pass
					if(int(voltage3_set) > 0):
						GPIO.output(RELAY_PIN, GPIO.HIGH)
						GPIO.output(RUN_LED_PIN, GPIO.HIGH)
						uart_send(voltage3_set, 1)
					else:
						uart_send(0, 0)
						GPIO.output(RELAY_PIN, GPIO.LOW)
						GPIO.output(RUN_LED_PIN, GPIO.LOW)
						
						for i in range(0,3):
							GPIO.output(BUZZER_PIN, GPIO.HIGH)
							sleep(1)
							GPIO.output(BUZZER_PIN, GPIO.LOW)
							sleep(1)

					stage2_is_running = 0
					stage3_is_running = 1

					time_s3()

			def time_s1():
				multirunscreen0_title_labelframe['bg']='grey70'
				voltage0_labelframe['bg'] = 'grey70'
				voltage0_label['bg']='grey70'
				timer0_labelframe['bg'] = 'grey70'
				m0_label['bg'] = 'grey70'
				twodot0_label['bg'] = 'grey70'
				s0_label['bg'] = 'grey70'

				multirunscreen1_title_labelframe['bg']='green yellow'
				voltage1_labelframe['bg'] = 'green yellow'
				voltage1_label['bg']='green yellow'
				timer1_labelframe['bg'] = 'green yellow'
				m1_label['bg'] = 'green yellow'
				twodot1_label['bg'] = 'green yellow'
				s1_label['bg'] = 'green yellow'

				global s1_set, m1_set, stage1_is_running, stage2_is_running, result_path, path_name, camera, solve_s1
				global current_measured

				s1_set = s1_set - 1
				if(s1_set<0):
					m1_set=m1_set-1
					s1_set=59
				s1_label.config(text = str('%02d'%s1_set))
				m1_label.config(text = str('%02d'%m1_set))

				adc = AnalogIn(ads, ADS.P2)
				adc_voltage = adc.voltage
				real_voltage = round(adc_voltage*max_voltage/divide_voltage)
				present_voltage_label.config(text="Voltage: " + str(real_voltage) + ' V')

				send_data = '\rSTATUS\r'
				ser.write(send_data.encode())
				if(ser.in_waiting>0):
					try:
						current_measured = float(ser.readline().decode('utf-8').rstrip())
					except:
						current_measured = 0.045
						pass

					if(current_measured < CURRENT_WARNING):
						present_current_label.config(text="LOW CURRENT WARNING !")
					else:
						present_current_label.config(text="")

				if(m1_set!=-1):
					solve_s1 = s1_label.after(1000, time_s1)
				else:
					m1_label.config(text = '00')
					s1_label.config(text = '00')
					try:
						s1_label.after_cancel(solve_s1)
					except:
						pass

					if(s1_raw!=0 or m1_raw!=0):
						try:
							camera_capture(path_name + 'stage1_result.png')
						except Exception as e:
							error = messagebox.showerror("ERR 03", str(e), icon = "error")
							if(error=='ok'):
								pass
					if(int(voltage2_set) > 0):
						GPIO.output(RELAY_PIN, GPIO.HIGH)
						GPIO.output(RUN_LED_PIN, GPIO.HIGH)
						uart_send(voltage2_set, 1)
					else:
						uart_send(0, 0)
						GPIO.output(RELAY_PIN, GPIO.LOW)
						GPIO.output(RUN_LED_PIN, GPIO.LOW)
						
						for i in range(0,3):
							GPIO.output(BUZZER_PIN, GPIO.HIGH)
							sleep(1)
							GPIO.output(BUZZER_PIN, GPIO.LOW)
							sleep(1)

					stage1_is_running = 0
					stage2_is_running = 1

					time_s2()

			def time_s0():
				global adc_value, s0_set, m0_set, stage0_is_running, stage1_is_running, result_path, path_name, camera, solve_s0
				global current_measured

				s0_set = s0_set - 1
				if(s0_set<0):
					m0_set=m0_set-1
					s0_set=59
				s0_label.config(text = str('%02d'%s0_set))
				m0_label.config(text = str('%02d'%m0_set))

				adc = AnalogIn(ads, ADS.P2)
				adc_voltage = adc.voltage
				real_voltage = round(adc_voltage*max_voltage/divide_voltage)
				present_voltage_label.config(text="Voltage: " + str(real_voltage) + ' V')

				send_data = '\rSTATUS\r'
				ser.write(send_data.encode())
				if(ser.in_waiting>0):
					try:
						current_measured = float(ser.readline().decode('utf-8').rstrip())
					except:
						current_measured = 0.045
						pass
					print("Data received:", current_measured)
					if(current_measured < CURRENT_WARNING):
						present_current_label.config(text="LOW CURRENT WARNING !")
					else:
						present_current_label.config(text="")

				if(m0_set!=-1 and running==1):
					solve_s0 = s0_label.after(1000, time_s0)
				else:
					m0_label.config(text = '00')
					s0_label.config(text = '00')
					s0_label.after_cancel(solve_s0)

					if(s0_raw!=0 or m0_raw!=0):
						try:
							camera_capture(path_name + 'stage0_result.png')
						except Exception as e:
							error = messagebox.showerror("ERR 03", str(e), icon = "error")
							if(error=='ok'):
								pass
					if(int(voltage1_set) > 0):
						GPIO.output(RELAY_PIN, GPIO.HIGH)
						GPIO.output(RUN_LED_PIN, GPIO.HIGH)
						uart_send(voltage1_set, 1)
					else:
						uart_send(0, 0)
						GPIO.output(RELAY_PIN, GPIO.LOW)
						GPIO.output(RUN_LED_PIN, GPIO.LOW)
						
						for i in range(0,3):
							GPIO.output(BUZZER_PIN, GPIO.HIGH)
							sleep(1)
							GPIO.output(BUZZER_PIN, GPIO.LOW)
							sleep(1) 

					stage0_is_running = 0
					stage1_is_running = 1

					time_s1()

			button_labelframe = LabelFrame(multirunscreen_labelframe, bg=FRAME_BACKGROUND_COLOR_1, width=622, height=66)
			button_labelframe.place(x=173,y=410)

			para_labelframe = LabelFrame(button_labelframe, bg=FRAME_BACKGROUND_COLOR_1, width=179, height=55)
			para_labelframe.place(x=3,y=4)
			present_voltage_label = Label(para_labelframe, font=('Helvetica',10,'bold'), fg='black', bg=FRAME_BACKGROUND_COLOR_1)
			present_voltage_label.place(x=5,y=5)
			present_current_label = Label(para_labelframe, font=('Helvetica',9,'bold'), fg='black', bg='red')
			present_current_label.place(x=5,y=28)

#                 def stop_click():
#                     global m0_set, s0_set
#                     m0_set = m0_raw
#                     s0_set = s0_set
#
#                     camera.close()
#                     root.destroy()

			def capture_click():
				global camera
				if(capture_button['text']=='CAPTURE'):
					if(stage0_is_running):
						s_cap = s0_raw - s0_set
						if(s_cap<0):
							s_cap = 60 + s_cap
							m_cap = m0_raw - m0_set -1
						else:
							m_cap = m0_raw - m0_set
						if(m_cap<0):
							m_cap=0

						output_result = str('stage0_' + '%02d'%m_cap) + ':' + str('%02d'%s_cap) +'.png'
					elif(stage1_is_running):
						s_cap = s1_raw - s1_set
						if(s_cap<0):
							s_cap = 60 + s_cap
							m_cap = m1_raw - m1_set -1
						else:
							m_cap = m1_raw - m1_set
						if(m_cap<0):
							m_cap=0

						output_result = str('stage1_' + '%02d'%m_cap) + ':' + str('%02d'%s_cap) +'.png'
					elif(stage2_is_running):
						s_cap = s2_raw - s2_set
						if(s_cap<0):
							s_cap = 60 + s_cap
							m_cap = m2_raw - m2_set -1
						else:
							m_cap = m2_raw - m2_set
						if(m_cap<0):
							m_cap=0
						output_result = str('stage2_' + '%02d'%m_cap) + ':' + str('%02d'%s_cap) +'.png'
					elif(stage3_is_running):
						s_cap = s1_raw - s1_set
						if(s_cap<0):
							s_cap = 60 + s_cap
							m_cap = m3_raw - m3_set -1
						else:
							m_cap = m3_raw - m3_set
						if(m_cap<0):
							m_cap=0
						output_result = str('stage3_' + '%02d'%m_cap) + ':' + str('%02d'%s_cap) +'.png'

					try:
						camera.capture(path_name + output_result)
						camera.stop_preview()
						msgbox = messagebox.showinfo('Capture Done','The picture have been saved !')
						if(msgbox=='ok'):
							camera_preview()

					except Exception as e:
						error = messagebox.showerror("ERR 03",str(e), icon = "error")
						if(error=='ok'):
							pass
				else:
					rsfile = filedialog.askopenfilename(initialdir = path_name, filetypes=[('jpg file','*.jpg')])
					if rsfile is not None:
						if(rsfile[len(rsfile)-3:]=='jpg'):
							a1 = Image.open(rsfile)
							#a1 = Image.open(path_name + 'result.jpg')
							crop_width, crop_height = a1.size
							scale_percent = 50
							width = int(crop_width * scale_percent / 100)
							height = int(crop_height * scale_percent / 100)
							display_img = a1.resize((width,height))
							a1_display = ImageTk.PhotoImage(display_img)
							a1_label = Label(preview_labelframe, image=a1_display)
							a1_label.image = a1_display
							a1_label.place(x=53,y=-5)
						else:
							pass

			def stop_click():
				global m0_set, s0_set, camera, solve_auto_capture, solve_s1, solve_s0, solve_s2, solve_s3
				try:
					camera.stop_preview()
				except:
					pass
				msgbox = messagebox.askquestion('STOP',' Do you want to go back to Main Screen ?', icon = 'question')
				if(msgbox=='yes'):
					uart_send(0, 0)
					GPIO.output(RUN_LED_PIN, GPIO.LOW)
					GPIO.output(RELAY_PIN, GPIO.LOW)

					if(stop_button['text'] != 'FINISH'):
						try:
							camera_capture(path_name + 'final_result.png')

							# ~ fr_coordinate = open('/home/pi/VE100/coordinates.txt')
							# ~ x_coordinate_17 = int(fr_coordinate.readline())
							# ~ x_coordinate_26 = int(fr_coordinate.readline().strip('\n'))
							# ~ y_coordinate = int(fr_coordinate.readline().strip('\n'))

							edit_img = Image.open(path_name + 'final_result.png')
							img = ImageDraw.Draw(edit_img)
							shape = [(1024, 768), (0,550)]
							img.rectangle(shape, fill ="lightgray", outline="lightgray")
							img_font_1_17 = ImageFont.truetype("/home/pi/VE100/arial.ttf", font_size_17)
							img_font_1_26 = ImageFont.truetype("/home/pi/VE100/arial.ttf", font_size_26)
							img_font_2 = ImageFont.truetype("/home/pi/VE100/arial.ttf", 23)
							if(num_well==17):
								x_coordinate = x_coordinate_17
								for i in range(0,9):
									img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_17, fill=(0,255,0))
									x_coordinate += well_distance_17

								x_coordinate = x_coordinate - well_distance_17 + pace_17
								for i in range(9,18):
									img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_17, fill=(0,255,0))
									x_coordinate += well_distance_17
							else:
								x_coordinate = x_coordinate_26
								for i in range(0,13):
									img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_26, fill=(0,255,0))
									x_coordinate += well_distance_26

								x_coordinate = x_coordinate - well_distance_26 + pace_26 
								for i in range(13,26):
									img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_26, fill=(0,255,0))
									# ~ if(i<18):
										# ~ x_coordinate += 26
									# ~ elif(i==18):
										# ~ x_coordinate += 30
									# ~ else:
									x_coordinate += well_distance_26

							if(num_well==26):
								r1 = 568
								r2 = 568
								r3 = 568
								r4 = 568
								r5 = 568
								for i in range(0,26):
									if(i<6):
										img.text((32,r1), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
										r1 += 32
									elif(i<12):
										img.text((232,r2), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
										r2 += 32
									elif(i<18):
										img.text((432,r3), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
										r3 += 32
									elif(i<24):
										img.text((632,r4), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
										r4 += 32
									else:
										img.text((832,r5), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
										r5 += 32
							else:
								r1 = 568
								r2 = 568
								r3 = 568
								r4 = 568
								r5 = 568
								for i in range(0,18):
									if(i<6):
										img.text((32,r1), str(i+1) + '. ' + str(well_name_list_17[i]), font=img_font_2, fill=(0,0,0))
										r1 += 32
									elif(i<12):
										img.text((232,r2), str(i+1) + '. ' + str(well_name_list_17[i]), font=img_font_2, fill=(0,0,0))
										r2 += 32
									else:
										img.text((432,r3), str(i+1) + '. ' + str(well_name_list_17[i]), font=img_font_2, fill=(0,0,0))
										r3 += 32
							edit_img.save(path_name + 'edit_img.png','png')

							# wb = Workbook()
	#                         sheet = wb.active
	#                         img = Img(path_name + 'edit_img.png')
	#                         img.height = 384
	#                         img.width = 512
	#                         img.anchor = 'B2'
	#                         sheet.add_image(img)

	#                         sheet.column_dimensions['J'].width = 18
	#                         sheet.column_dimensions['K'].width = 18

	#                         if(num_well==17):
	#                             pass
	#                         else:
	#                             for i in range(0,13):
	#                                 sheet['J'+ str(i+2)] = str(i+1) + '. ' + str(well_name_list_26[i])
	#                                 sheet['K'+ str(i+2)] = str(i+14) + '. ' + str(well_name_list_26[i+13])

	#                         wb.save(path_name + 'result.xlsx')

						except Exception as e:
							error = messagebox.showerror("ERR 03", str(e), icon = "error")
							if(error=='ok'):
								pass

						if(account_active and automail2_is_on):
							msgbox = messagebox.askquestion('STOP','Do you want to send email ?', icon = 'question')
							if(msgbox=='yes'):
								shutil.make_archive(path_name, format='zip', root_dir = path_name)
								try:
									sendmail(recipient_email, folder_name_set , 'This is an email from EV100 device.', 
									 + '/' + folder_name_set + '.zip')
								except Exception as e:
									try:
										camera.stop_preview()
									except:
										pass
									error = messagebox.showerror("ERR 02", str(e), icon = "error")
									if(error=='ok'):
										pass

					GPIO.output(BLUE_LIGHT_PIN, GPIO.LOW)

					m0_set = m0_raw
					s0_set = s0_raw

					try:
						s0_label.after_cancel(solve_s0)
					except:
						pass
					try:
						s1_label.after_cancel(solve_s1)
					except:
						pass
					try:
						s2_label.after_cancel(solve_s2)
					except:
						pass
					try:
						s3_label.after_cancel(solve_s3)
					except:
						pass
					try:
						root.after_cancel(solve_auto_capture)
					except:
						pass
					global running
					running = 0
					multirunscreen_labelframe.destroy()
					main()
				else:
					if(running==1):
						GPIO.output(RUN_LED_PIN, GPIO.HIGH)
						GPIO.output(BLUE_LIGHT_PIN, GPIO.HIGH)
						GPIO.output(RELAY_PIN, GPIO.HIGH)
						camera_preview()
			capture_button = Button(button_labelframe, bg=BUTTON_BACKGROUND_COLOR_3, bd=1, text='CAPTURE',font=("Courier",13,'bold'), width=16, height=2, command=capture_click)
			capture_button.place(x=210,y=3)
			stop_button = Button(button_labelframe, bg=BUTTON_BACKGROUND_COLOR_4, bd=1, text='STOP',font=("Courier",13,'bold'), width=16, height=2, command =stop_click)
			stop_button.place(x=427,y=3)

			GPIO.output(RUN_LED_PIN, GPIO.HIGH)
			GPIO.output(BLUE_LIGHT_PIN, GPIO.HIGH)
			GPIO.output(RELAY_PIN, GPIO.HIGH)

			camera_preview()
			if(voltage0_set != 0):
				uart_send(voltage0_set,1)
				time_s0()
			else:
				uart_send(voltage1_set,1)
				time_s1()

	def back_click():
		multirunscreen_labelframe.destroy()
		main()

	def save_click():
		global voltage1_set, voltage2_set, voltage3_set, m1_set, s1_set, m2_set, s2_set, m3_set, s3_set, m0_raw, s0_raw, m1_raw, s1_raw, m2_raw, s2_raw, m3_raw, s3_raw, recipient_email, folder_name, path_name
		msg = messagebox.askquestion("SAVE", "Do you want to save ?")
		if(msg=='yes'):
			if(voltage1_entry.get()=='' or voltage2_entry.get()=='' or voltage3_entry.get()==''):
				messagebox.showwarning("","Please enter the voltage !")
			elif(m1_entry.get()=='' or m2_entry.get()=='' or m3_entry.get()=='' or
					s1_entry.get()=='' or s2_entry.get()=='' or s3_entry.get()==''):
				messagebox.showwarning("","Please enter the timer !")
			else:
				voltage1_set = voltage1_entry.get()
				voltage2_set = voltage2_entry.get()
				voltage3_set = voltage3_entry.get()
				m1_set = int(m1_entry.get())
				m2_set = int(m2_entry.get())
				m3_set = int(m3_entry.get())
				s1_set = int(s1_entry.get())
				s2_set = int(s2_entry.get())
				s3_set = int(s3_entry.get())

				fw = open('/home/pi/VE100/parameters2.txt','w')
				fw.writelines(str(voltage1_set) + '\n')
				fw.writelines(str('%02d'%m1_set) + str('%02d'%s1_set) +  '\n')
				fw.writelines(str(voltage2_set) + '\n')
				fw.writelines(str('%02d'%m2_set) + str('%02d'%s2_set) + '\n')
				fw.writelines(str(voltage3_set) + '\n')
				fw.writelines(str('%02d'%m3_set) + str('%02d'%s3_set) + '\n')
				fw.writelines(str(automail2_is_on) + '\n')
				messagebox.showinfo("SAVE", 'Saved !')

	back_button = Button(multirunscreen_labelframe, bd=1, bg=BUTTON_BACKGROUND_COLOR_1, font=('Courier',12, 'bold'), text='BACK', width=8, height=3, command=back_click)
	back_button.place(x=0,y=405)
	save_button = Button(multirunscreen_labelframe, bg=BUTTON_BACKGROUND_COLOR_1, font=('Courier',12, 'bold'), bd=1,text='SAVE', width=8, height=3, command=save_click)
	save_button.place(x=109,y=405)
	run1_button = Button(multirunscreen_labelframe, bg=BUTTON_BACKGROUND_COLOR_1, font=('Courier',12, 'bold'), bd=1,text='RUN', width=8, height=3, command=run1_click)
	run1_button.place(x=687,y=405)

def oneStepRunScreen():
	global automail1_is_on
	global current_measured

	fr_para1 = open("/home/pi/VE100/parameters1.txt","r")
	voltage_set = int(fr_para1.readline())
	timer_set = int(fr_para1.readline().strip('\n'))
	auto_capture_timer_set = int(fr_para1.readline().strip('\n'))
	automail1_is_on = int(fr_para1.readline().strip('\n'))
	m_set = round(timer_set/100)
	s_set = timer_set - (m_set*100)

	onerunscreen_labelframe = LabelFrame(root, bg=FRAME_BACKGROUND_COLOR_1, width=800, height=480)
	onerunscreen_labelframe.place(x=0,y=0)
	onerunscreen_title_label =  Label(onerunscreen_labelframe, text="SETTINGS", font=("Courier",14, "bold"), fg=LABEL_TEXT_COLOR_1, bg=LABEL_BACKGROUND_COLOR_1, width=72, height=2)
	onerunscreen_title_label.place(x=0,y=1)

	setting1_labelframe = LabelFrame(onerunscreen_labelframe, bg=FRAME_BACKGROUND_COLOR_1, width=263, height=355)
	setting1_labelframe.place(x=0,y=51)
	setting2_labelframe = LabelFrame(onerunscreen_labelframe, bg=FRAME_BACKGROUND_COLOR_1, width=263, height=355)
	setting2_labelframe.place(x=266,y=51)
	setting3_labelframe = LabelFrame(onerunscreen_labelframe, bg=FRAME_BACKGROUND_COLOR_1, width=264, height=355)
	setting3_labelframe.place(x=532,y=51)

	voltage_labelframe = LabelFrame(setting2_labelframe, bg=FRAME_BACKGROUND_COLOR_2, fg=FRAME_TEXT_COLOR_2, text='VOLTAGE (VDC)', font=('Courier',12), width=238, height=100)
	voltage_labelframe.place(x=10,y=16)
	timer_labelframe = LabelFrame(setting2_labelframe, bg=FRAME_BACKGROUND_COLOR_2, fg=FRAME_TEXT_COLOR_2, text='RUN TIMER (min:sec)', font=('Courier',12), width=238, height=100)
	timer_labelframe.place(x=10,y=124)
	autocap_labelframe = LabelFrame(setting2_labelframe, bg=FRAME_BACKGROUND_COLOR_2, fg=FRAME_TEXT_COLOR_2, text='CAPTURE TIMER (min)', font=('Courier',12), width=238, height=100)
	autocap_labelframe.place(x=10,y=232)
	twodot_label = Label(timer_labelframe, fg=LABEL_TEXT_COLOR_3, bg=FRAME_BACKGROUND_COLOR_2, text =':', font=('Courier', 20))
	twodot_label.place(x=106, y=15)

	voltage_entry = Entry(voltage_labelframe, fg='red', justify='center',width=3, font=('Courier',25))
	voltage_entry.place(x=84,y=12)
	voltage_entry.insert(0, voltage_set)
	m_entry = Entry(timer_labelframe, fg='red', justify='center',width=3, font=('Courier',25))
	m_entry.place(x=31,y=12)
	m_entry.insert(0, str('%02d'%m_set))
	s_entry = Entry(timer_labelframe, fg='red', justify='center',width=3, font=('Courier',25))
	s_entry.place(x=135,y=12)
	s_entry.insert(0, str('%02d'%s_set))
	autocap_entry = Entry(autocap_labelframe, fg='red', justify='center',width=3, font=('Courier',25))
	autocap_entry.place(x=84,y=12)
	autocap_entry.insert(0, auto_capture_timer_set)

	automail_labelframe = LabelFrame(onerunscreen_labelframe, bg=FRAME_BACKGROUND_COLOR_2, text='AUTO SEND EMAIL', width=467, height=68)
	automail_labelframe.place(x=219,y=407)

	mail_label = Label(automail_labelframe, bg=FRAME_BACKGROUND_COLOR_2, text='To:', font=('Courier, 15'))
	mail_label.place(x=130, y=7)
	mail_entry = Entry(automail_labelframe, justify='right',width=28, font=('Courier',13))
	mail_entry.place(x=171,y=7)
	mail_entry.insert(0,autofill_email)

	def automail_on_click():
		global automail1_is_on
		automail1_is_on = 1
		automail_on_button['bg'] = 'lawn green'
		automail_off_button['bg'] = MAINMENU_BUTTON_DISABLE_COLOR
		mail_entry['state'] = 'normal'
		if(mail_entry.get()==''):
			mail_entry.insert(0,autofill_email)
		mail_label['fg']='black'
	def automail_off_click():
		global automail1_is_on
		automail1_is_on = 0
		automail_on_button['bg'] = MAINMENU_BUTTON_DISABLE_COLOR
		automail_off_button['bg'] = 'red'
		mail_entry.delete(0,END)
		mail_entry['state'] = 'disable'
		mail_label['fg']='grey75'

	automail_on_button = Button(automail_labelframe,bd=0,text='ON', width=4, height=2, command=automail_on_click)
	automail_on_button.place(x=5,y=-1)
	automail_off_button = Button(automail_labelframe, bg=MAINMENU_BUTTON_DISABLE_COLOR,bd=0,text='OFF', width=4, height=2, command=automail_off_click)
	automail_off_button.place(x=63,y=-1)
	if(account_active==0):
		automail_on_button['state'] = 'disable'
		automail_off_button['state'] = 'disable'

	if(automail1_is_on):
		automail_on_click()
	else:
		automail_off_click()

	def run1_click():
		global auto_capture_timer_set ,voltage_set, m_set, s_set, m0_set, s0_set ,m_raw, s_raw, m0_raw, s0_raw, recipient_email, folder_name, path_name
		global folder_name_set, path_name
		
		if(voltage_entry.get()==''):
			messagebox.showwarning("","Please enter the voltage !")
		elif((voltage_entry.get().isnumeric())==0 ):
			messagebox.showwarning("","Voltage must be between 12 and 80 VDC")
		elif(int(voltage_entry.get())<12 or int(voltage_entry.get())>80):
			messagebox.showwarning("","Voltage must be between 12 and 80 VDC")
		elif(m_entry.get()=='' or s_entry.get()==''):
			messagebox.showwarning("","Please enter the run timer !")
		elif(autocap_entry.get()==''):
			messagebox.showwarning("","Please enter the auto capture timer !")
		elif((mail_entry.get()=='' or mail_entry.get()=='@gmail.com') and automail1_is_on):
			messagebox.showwarning("","Please enter the recipient email !")
		else:
			global running
			running = 1
			
			global autofill_email
			if(automail1_is_on):
				recipient_email = mail_entry.get()
				fw_email = open('/home/pi/VE100/.oldemail.txt', 'w')
				fw_email.writelines(recipient_email + '\n')
				fw_email.close()
				autofill_email = recipient_email

			voltage_set = voltage_entry.get()
			m_set = int(m_entry.get())
			s_set = int(s_entry.get())
			auto_capture_timer_set = int(autocap_entry.get())
			m_raw = m_set
			s_raw = s_set
			m0_raw = m0_set
			s0_raw = s0_set

			subprocess.call(["scrot",path_name + 'parameters.png'])

			multirunscreen_labelframe = LabelFrame(root, bg='grey85', width=800, height=480)
			multirunscreen_labelframe.place(x=0,y=0)
			multirunscreen0_title_labelframe = LabelFrame(multirunscreen_labelframe, text="DEFAULT STAGE", font=('Courier', 13, 'bold'), bg='green yellow', width=170, height=118)
			multirunscreen0_title_labelframe.place(x=0,y=1)
			multirunscreen1_title_labelframe = LabelFrame(multirunscreen_labelframe, text="STAGE 1", font=('Courier', 13, 'bold'), bg='grey70', width=170, height=118)
			multirunscreen1_title_labelframe.place(x=0,y=120)

			voltage0_labelframe = LabelFrame(multirunscreen0_title_labelframe, text="Voltage", font=('Courier', 12), bg='green yellow', width=140, height=45)
			voltage0_labelframe.place(x=13,y=-5)
			timer0_labelframe = LabelFrame(multirunscreen0_title_labelframe, text="Time Left", font=('Courier', 12), bg='green yellow', width=140, height=45)
			timer0_labelframe.place(x=13,y=42)
			voltage0_label = Label(voltage0_labelframe, fg ='red2', text=str(voltage0_set) + ' VDC', font=('calibri',14,'bold'), bg='green yellow')
			voltage0_label.place(x=36, y=-2)
			m0_label = Label(timer0_labelframe, fg ='red2', text=str('%.02d'%m0_set), font=('calibri',14,'bold'), bg='green yellow')
			m0_label.place(x=35, y=-2)
			twodot0_label = Label(timer0_labelframe, fg ='red2', text=':', font=('calibri',14,'bold'), bg='green yellow')
			twodot0_label.place(x=63,y=-2)
			s0_label = Label(timer0_labelframe, fg ='red2', text=str('%.02d'%s0_set), font=('calibri',14,'bold'), bg='green yellow')
			s0_label.place(x=77, y=-2)

			voltage1_labelframe = LabelFrame(multirunscreen1_title_labelframe, text="Voltage", font=('Courier', 12), bg='grey70', width=140, height=45)
			voltage1_labelframe.place(x=13,y=-5)
			timer1_labelframe = LabelFrame(multirunscreen1_title_labelframe, text="Time Left", font=('Courier', 12), bg='grey70', width=140, height=45)
			timer1_labelframe.place(x=13,y=42)
			voltage1_label = Label(voltage1_labelframe, fg ='red2', text=str(voltage_set) + ' VDC', font=('calibri',14,'bold'), bg='grey70')
			voltage1_label.place(x=36, y=-2)
			m1_label = Label(timer1_labelframe, fg ='red2', text=str('%.02d'%m_set), font=('calibri',14,'bold'), bg='grey70')
			m1_label.place(x=35, y=-2)
			twodot1_label = Label(timer1_labelframe, fg ='red2', text=':', font=('calibri',14,'bold'), bg='grey70')
			twodot1_label.place(x=63,y=-2)
			s1_label = Label(timer1_labelframe, fg ='red2', text=str('%.02d'%s_set), font=('calibri',14,'bold'), bg='grey70')
			s1_label.place(x=77, y=-2)

			preview_labelframe = LabelFrame(multirunscreen_labelframe, bg='black', fg='red', text="◉ MONITOR", font=("Courier",13,'bold'), width=622, height=410)
			preview_labelframe.place(x=173,y=0)


			limit1_canvas = Canvas(preview_labelframe, bg='grey70', bd=0, width=54, height=376)
			limit1_canvas.place(x=0,y=1)
			limit2_canvas = Canvas(preview_labelframe, bg='grey70', bd=0, width=54, height=376)
			limit2_canvas.place(x=562,y=1)
			limit_line1 = limit1_canvas.create_line(x0_limit, y0_limit, x1_limit, y1_limit, fill='red', width=2)
			limit_line2 = limit2_canvas.create_line(x0_limit, y0_limit, x1_limit, y1_limit, fill='red', width=2)

			t_progressbar = atk.RadialProgressbar(preview_labelframe, fg='cyan')
			t_progressbar.place(x=260,y=130)
			t_progressbar.start()
			tprocess_label = Label(preview_labelframe, bg='black', fg='white smoke', text='Processing\r...', font=("Courier",9,'bold'))
			tprocess_label.place(x=272,y=172)

			def time_s1():
				multirunscreen0_title_labelframe['bg']='grey70'
				voltage0_labelframe['bg'] = 'grey70'
				voltage0_label['bg']='grey70'
				timer0_labelframe['bg'] = 'grey70'
				m0_label['bg'] = 'grey70'
				twodot0_label['bg'] = 'grey70'
				s0_label['bg'] = 'grey70'

				multirunscreen1_title_labelframe['bg']='green yellow'
				voltage1_labelframe['bg'] = 'green yellow'
				voltage1_label['bg']='green yellow'
				timer1_labelframe['bg'] = 'green yellow'
				m1_label['bg'] = 'green yellow'
				twodot1_label['bg'] = 'green yellow'
				s1_label['bg'] = 'green yellow'

				global final_cap, s_set, m_set, s0_set, m0_set, stage0_is_running, stage1_is_running, result_path, path_name, camera, solve_s1, solve_auto_capture
				global current_measured
				global x_coordinate_17, x_coordinate_26, y_coordinate

				s_set = s_set - 1
				if(s_set<0):
					m_set=m_set-1
					s_set=59
				s1_label.config(text = str('%02d'%s_set))
				m1_label.config(text = str('%02d'%m_set))

				adc = AnalogIn(ads, ADS.P2)
				adc_voltage = adc.voltage
				real_voltage = round(adc_voltage*max_voltage/divide_voltage)
				present_voltage_label.config(text="Voltage: " + str(real_voltage) + ' V')

				send_data = '\rSTATUS\r'
				ser.write(send_data.encode())
				if(ser.in_waiting>0):
					try:
						current_measured = float(ser.readline().decode('utf-8').rstrip())
					except:
						current_measured = 0.045
						pass
					print("Data received:", current_measured)
					if(current_measured < CURRENT_WARNING):
						present_current_label.config(text="LOW CURRENT WARNING !")
					else:
						present_current_label.config(text="")

				if(m_set!=-1):
					solve_s1 = s1_label.after(1000, time_s1)
				else:
					limit1_canvas.place_forget()  
					limit2_canvas.place_forget()
					present_voltage_label.config(text="Voltage: 0 V")
					present_current_label.config(text="")

					GPIO.output(RUN_LED_PIN, GPIO.LOW)
					GPIO.output(RELAY_PIN, GPIO.LOW)
					uart_send(0,0)

					m1_label.config(text = '00')
					s1_label.config(text = '00')
					try:
						s1_label.after_cancel(solve_s1)
					except:
						pass

					final_cap = 1

					try:
						camera_capture(path_name + 'stage1_result.png')

						# ~ fr_coordinate = open('/home/pi/VE100/coordinates.txt')
						# ~ x_coordinate_17 = int(fr_coordinate.readline())
						# ~ x_coordinate_26 = int(fr_coordinate.readline().strip('\n'))
						# ~ y_coordinate =  int(fr_coordinate.readline().strip('\n'))

						edit_img = Image.open(path_name + 'stage1_result.png')
						img = ImageDraw.Draw(edit_img)
						shape = [(1024, 768), (0,550)]
						img.rectangle(shape, fill ="lightgray", outline="lightgray")
						img_font_1_17 = ImageFont.truetype("/home/pi/VE100/arial.ttf", font_size_17)
						img_font_1_26 = ImageFont.truetype("/home/pi/VE100/arial.ttf", font_size_26)
						img_font_2 = ImageFont.truetype("/home/pi/VE100/arial.ttf", 23)
						if(num_well==17):
							x_coordinate = x_coordinate_17
							for i in range(0,9):
								img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_17, fill=(0,255,0))
								x_coordinate += well_distance_17

							x_coordinate = x_coordinate - well_distance_17 + pace_17
							for i in range(9,18):
								img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_17, fill=(0,255,0))
								x_coordinate += well_distance_17
						else:
							x_coordinate = x_coordinate_26
							for i in range(0,13):
								img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_26, fill=(0,255,0))
								x_coordinate += well_distance_26

							x_coordinate = x_coordinate - well_distance_26 + pace_26
							for i in range(13,26):
								img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_26, fill=(0,255,0))
								# ~ if(i<18):
									# ~ x_coordinate += 26
								# ~ elif(i==18):
									# ~ x_coordinate += 30
								# ~ else:
								x_coordinate += well_distance_26

						if(num_well==26):
							r1 = 568
							r2 = 568
							r3 = 568
							r4 = 568
							r5 = 568
							for i in range(0,26):
								if(i<6):
									img.text((32,r1), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
									r1 += 32
								elif(i<12):
									img.text((232,r2), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
									r2 += 32
								elif(i<18):
									img.text((432,r3), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
									r3 += 32
								elif(i<24):
									img.text((632,r4), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
									r4 += 32
								else:
									img.text((832,r5), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
									r5 += 32
						else:
							r1 = 568
							r2 = 568
							r3 = 568
							r4 = 568
							r5 = 568
							for i in range(0,18):
								if(i<6):
									img.text((32,r1), str(i+1) + '. ' + str(well_name_list_17[i]), font=img_font_2, fill=(0,0,0))
									r1 += 32
								elif(i<12):
									img.text((232,r2), str(i+1) + '. ' + str(well_name_list_17[i]), font=img_font_2, fill=(0,0,0))
									r2 += 32
								else:
									img.text((432,r3), str(i+1) + '. ' + str(well_name_list_17[i]), font=img_font_2, fill=(0,0,0))
									r3 += 32

						edit_img.save(path_name + 'edit_img.png','png')

						# wb = Workbook()
#                         sheet = wb.active
#                         img = Img(path_name + 'edit_img.png')
#                         img.height = 384
#                         img.width = 512
#                         img.anchor = 'B2'
#                         sheet.add_image(img)

#                         sheet.column_dimensions['J'].width = 18
#                         sheet.column_dimensions['K'].width = 18

#                         if(num_well==17):
#                             pass
#                         else:
#                             for i in range(0,13):
#                                 sheet['J'+ str(i+2)] = str(i+1) + '. ' + str(well_name_list_26[i])
#                                 sheet['K'+ str(i+2)] = str(i+14) + '. ' + str(well_name_list_26[i+13])

#                         wb.save(path_name + 'result.xlsx')

					except Exception as e:
						error = messagebox.showerror("ERR 03", str(e), icon = "error")
						if(error=='ok'):
							pass

					m0_set = m0_raw
					s0_set = s0_raw

					multirunscreen1_title_labelframe['bg']='grey70'
					voltage1_labelframe['bg'] = 'grey70'
					voltage1_label['bg']='grey70'
					timer1_labelframe['bg'] = 'grey70'
					m1_label['bg'] = 'grey70'
					twodot1_label['bg'] = 'grey70'
					s1_label['bg'] = 'grey70'

					if(account_active and automail1_is_on):
						shutil.make_archive(path_name, format='zip', root_dir = path_name)
						try:
							sendmail(recipient_email, folder_name_set , 'This is an email from EV100 device.', path_name_0 + '/' + folder_name_set + '.zip')
						except Exception as e:
							try:
								camera.stop_preview()
							except:
								pass
							error = messagebox.showerror("ERR 02", str(e), icon = "error")
							if(error=='ok'):
								pass
					t_progressbar.place_forget()
					tprocess_label['font'] = ('Courier',20, 'bold')
					tprocess_label['text'] = 'COMPLETED'
					tprocess_label['fg'] = 'lawn green'
					tprocess_label.place(x=236,y=160)

					capture_button['text'] = 'VIEW RESULT'
					stop_button['text'] = 'FINISH'

					stage1=is_running = 0
					stage0_is_running = 1

					GPIO.output(RELAY_PIN, GPIO.LOW)
					GPIO.output(BLUE_LIGHT_PIN, GPIO.LOW)
					GPIO.output(RUN_LED_PIN, GPIO.LOW)
					
					for i in range(0,3):
						GPIO.output(BUZZER_PIN, GPIO.HIGH)
						sleep(0.5)
						GPIO.output(BUZZER_PIN, GPIO.LOW)
						sleep(0.5)
							
					root.after_cancel(solve_auto_capture)
					try:
						camera.stop_preview()
					except:
						pass

					global running
					running = 0

			def time_s0():
				global s0_set, m0_set, stage0_is_running, stage1_is_running, result_path, path_name, camera, solve_s0
				global current_measured

				s0_set = s0_set - 1
				if(s0_set<0):
					m0_set=m0_set-1
					s0_set=59
				s0_label.config(text = str('%02d'%s0_set))
				m0_label.config(text = str('%02d'%m0_set))

				adc = AnalogIn(ads, ADS.P2)
				adc_voltage = adc.voltage
				real_voltage = round(adc_voltage*max_voltage/divide_voltage)
				present_voltage_label.config(text="Voltage: " + str(real_voltage) + ' V')

				send_data = '\rSTATUS\r'
				ser.write(send_data.encode())
				if(ser.in_waiting>0):
					try:
						current_measured = float(ser.readline().decode('utf-8').rstrip())
					except:
						current_measured = 0.045
						pass
					print("Data received:", current_measured)
					if(current_measured < CURRENT_WARNING):
						present_current_label.config(text="LOW CURRENT WARNING !")
					else:
						present_current_label.config(text="")

				if(m0_set!=-1):
					solve_s0 = s0_label.after(1000, time_s0)
				else:
					m0_label.config(text = '00')
					s0_label.config(text = '00')
					s0_label.after_cancel(solve_s0)

					if(s0_raw!=0 or m0_raw!=0):
						try:
							camera_capture(path_name + 'stage0_result.png')
						except Exception as e:
							error = messagebox.showerror("ERR 03", str(e), icon = "error")
							if(error=='ok'):
								pass
					if(int(voltage_set) > 0):
						GPIO.output(RELAY_PIN, GPIO.HIGH)
						GPIO.output(RUN_LED_PIN, GPIO.HIGH)
						uart_send(voltage_set, 1)
					else:
						uart_send(0, 0)
						GPIO.output(RELAY_PIN, GPIO.LOW)
						GPIO.output(RUN_LED_PIN, GPIO.LOW)
						
						for i in range(0,3):
							GPIO.output(BUZZER_PIN, GPIO.HIGH)
							sleep(1)
							GPIO.output(BUZZER_PIN, GPIO.LOW)
							sleep(1)

					stage0_is_running = 0
					stage1_is_running = 1

					time_s1()

			button_labelframe = LabelFrame(multirunscreen_labelframe, bg='grey90', width=622, height=66)
			button_labelframe.place(x=173,y=410)

			para_labelframe = LabelFrame(button_labelframe, bg=FRAME_BACKGROUND_COLOR_1, width=179, height=55)
			para_labelframe.place(x=3,y=4)
			present_voltage_label = Label(para_labelframe, font=('Helvetica',10,'bold'), fg='black', bg=FRAME_BACKGROUND_COLOR_1)
			present_voltage_label.place(x=5,y=5)
			present_current_label = Label(para_labelframe, font=('Helvetica',9,'bold'), fg='black', bg='red')
			present_current_label.place(x=5,y=28)

			def stop_click():
				global m0_set, s0_set, camera, solve_auto_capture, solve_s1, solve_s0
				try:
					camera.stop_preview()
				except:
					pass
				msgbox = messagebox.askquestion('STOP',' Are you sure you want to stop running ?', icon = 'question')
				if(msgbox=='yes'):
					uart_send(0, 0)
					GPIO.output(BLUE_LIGHT_PIN, GPIO.LOW)
					GPIO.output(RELAY_PIN, GPIO.LOW)

					if(stop_button['text'] != 'FINISH'):
						try:
							camera_capture(path_name + 'final_result.png')

							# ~ fr_coordinate = open('/home/pi/VE100/coordinates.txt')
							# ~ x_coordinate_17 = int(fr_coordinate.readline())
							# ~ x_coordinate_26 = int(fr_coordinate.readline().strip('\n'))
							# ~ y_coordinate = int(fr_coordinate.readline().strip('\n'))

							edit_img = Image.open(path_name + 'final_result.png')
							img = ImageDraw.Draw(edit_img)
							shape = [(1024, 768), (0,550)]
							img.rectangle(shape, fill ="lightgray", outline="lightgray")
							img_font_1_17 = ImageFont.truetype("/home/pi/VE100/arial.ttf", font_size_17)
							img_font_1_26 = ImageFont.truetype("/home/pi/VE100/arial.ttf", font_size_26)
							img_font_2 = ImageFont.truetype("/home/pi/VE100/arial.ttf", 23)
							if(num_well==17):
								x_coordinate = x_coordinate_17
								for i in range(0,9):
									img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_17, fill=(0,255,0))
									x_coordinate += well_distance_17

								x_coordinate = x_coordinate - well_distance_17 + pace_17
								for i in range(9,18):
									img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_17, fill=(0,255,0))
									x_coordinate += well_distance_17
							else:
								x_coordinate = x_coordinate_26
								for i in range(0,13):
									img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_26, fill=(0,255,0))
									x_coordinate += well_distance_26

								x_coordinate = x_coordinate - well_distance_26 + pace_26
								for i in range(13,26):
									img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_26, fill=(0,255,0))
									# ~ if(i<18):
										# ~ x_coordinate += 26
									# ~ elif(i==18):
										# ~ x_coordinate += 30
									# ~ else:
									x_coordinate += well_distance_26

							if(num_well==26):
								r1 = 568
								r2 = 568
								r3 = 568
								r4 = 568
								r5 = 568
								for i in range(0,26):
									if(i<6):
										img.text((32,r1), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
										r1 += 32
									elif(i<12):
										img.text((232,r2), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
										r2 += 32
									elif(i<18):
										img.text((432,r3), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
										r3 += 32
									elif(i<24):
										img.text((632,r4), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
										r4 += 32
									else:
										img.text((832,r5), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
										r5 += 32
							else:
								r1 = 568
								r2 = 568
								r3 = 568
								r4 = 568
								r5 = 568
								for i in range(0,18):
									if(i<6):
										img.text((32,r1), str(i+1) + '. ' + str(well_name_list_17[i]), font=img_font_2, fill=(0,0,0))
										r1 += 32
									elif(i<12):
										img.text((232,r2), str(i+1) + '. ' + str(well_name_list_17[i]), font=img_font_2, fill=(0,0,0))
										r2 += 32
									else:
										img.text((432,r3), str(i+1) + '. ' + str(well_name_list_17[i]), font=img_font_2, fill=(0,0,0))
										r3 += 32
							edit_img.save(path_name + 'edit_img.png','png')

							# wb = Workbook()
	#                         sheet = wb.active
	#                         img = Img(path_name + 'edit_img.png')
	#                         img.height = 384
	#                         img.width = 512
	#                         img.anchor = 'B2'
	#                         sheet.add_image(img)

	#                         sheet.column_dimensions['J'].width = 18
	#                         sheet.column_dimensions['K'].width = 18

	#                         if(num_well==17):
	#                             pass
	#                         else:
	#                             for i in range(0,13):
	#                                 sheet['J'+ str(i+2)] = str(i+1) + '. ' + str(well_name_list_26[i])
	#                                 sheet['K'+ str(i+2)] = str(i+14) + '. ' + str(well_name_list_26[i+13])

	#                         wb.save(path_name + 'result.xlsx')

						except Exception as e:
							error = messagebox.showerror("ERR 03", str(e), icon = "error")
							if(error=='ok'):
								pass

						if(account_active and automail2_is_on):
							msgbox = messagebox.askquestion('STOP','Do you want to send email ?', icon = 'question')
							if(msgbox=='yes'):
								shutil.make_archive(path_name, format='zip', root_dir = path_name)
								try:
									sendmail(recipient_email, folder_name_set, 'This is an email from EV100 device.', path_name_0 + '/' + folder_name_set + '.zip')
								except Exception as e:
									try:
										camera.stop_preview()
									except:
										pass
									error = messagebox.showerror("ERR 02", str(e), icon = "error")
									if(error=='ok'):
										pass

					GPIO.output(RUN_LED_PIN, GPIO.LOW)

					m0_set = m0_raw
					s0_set = s0_raw

					try:
						s0_label.after_cancel(solve_s0)
					except:
						pass
					try:
						s1_label.after_cancel(solve_s1)
					except:
						pass
					try:
						root.after_cancel(solve_auto_capture)
					except:
						pass
					global running
					running = 0
					onerunscreen_labelframe.destroy()
					main()
				else:
					if(running==1):
						GPIO.output(RUN_LED_PIN, GPIO.HIGH)
						GPIO.output(BLUE_LIGHT_PIN, GPIO.HIGH)
						GPIO.output(RELAY_PIN, GPIO.HIGH)
						camera_preview()
			
			def capture_click():
				global camera

				if(capture_button['text']=='CAPTURE'):
					if(stage0_is_running):
						s_cap = s0_raw - s0_set
						if(s_cap<0):
							s_cap = 60 + s_cap
							m_cap = m0_raw - m0_set -1
						else:
							m_cap = m0_raw - m0_set
						if(m_cap<0):
							m_cap=0
							
						output_result = str('stage0_' + '%02d'%m_cap) + ':' + str('%02d'%s_cap) +'.jpg'
					elif(stage1_is_running):
						s_cap = s1_raw - s1_set
						if(s_cap<0):
							s_cap = 60 + s_cap
							m_cap = m1_raw - m1_set -1
						else:
							m_cap = m1_raw - m1_set
						if(m_cap<0):
							m_cap=0

						output_result = str('stage1_' + '%02d'%m_cap) + ':' + str('%02d'%s_cap) +'.jpg'

					try:
						camera.capture(path_name + output_result)
						camera.stop_preview()
						msgbox = messagebox.showinfo('Capture Done','The picture have been saved !')
						if(msgbox=='ok'):
							camera_preview()

					except Exception as e:
						error = messagebox.showerror("ERR 03", str(e), icon = "error")
						if(error=='ok'):
							pass
				else:

					rsfile = filedialog.askopenfilename(initialdir = path_name, filetypes=[('jpg file','*.jpg')])
					if rsfile is not None:
						if(rsfile[len(rsfile)-3:]=='jpg'):
							a1 = Image.open(rsfile)
							#a1 = Image.open(path_name + 'result.jpg')
							crop_width, crop_height = a1.size
							scale_percent = 50
							width = int(crop_width * scale_percent / 100)
							height = int(crop_height * scale_percent / 100)
							display_img = a1.resize((width,height))
							a1_display = ImageTk.PhotoImage(display_img)
							a1_label = Label(preview_labelframe, image=a1_display)
							a1_label.image = a1_display
							a1_label.place(x=53,y=-5)
						else:
								pass

			capture_button = Button(button_labelframe, bg=BUTTON_BACKGROUND_COLOR_3, bd=1, text='CAPTURE',font=("Courier",13,'bold'), width=16, height=2, command=capture_click)
			capture_button.place(x=210,y=3)
			stop_button = Button(button_labelframe, bg=BUTTON_BACKGROUND_COLOR_4, bd=1, text='STOP',font=("Courier",13,'bold'), width=16, height=2, command =stop_click)
			stop_button.place(x=427,y=3)

			GPIO.output(BLUE_LIGHT_PIN, GPIO.HIGH)
			GPIO.output(RELAY_PIN, GPIO.HIGH)
			GPIO.output(RUN_LED_PIN, GPIO.HIGH)

			camera_preview()
			if(voltage0_set != 0):
				uart_send(voltage0_set,1)
				time_s0()
			else:
				uart_send(voltage_set,1)
				time_s1()

			def auto_capture():
				global solve_auto_capture, count, timer_set, auto_capture_call, path_name, time_capture
				if(auto_capture_timer_set !=1 or auto_capture_call !=0):
					output = 'auto_cap_' + str(count) +'.jpg'
					try:
						camera_capture(path_name + output)
					except Exception as e:
						error = messagebox.showerror("ERR 03", str(e), icon = "error")
						if(error=='ok'):
							pass
					count += 1
				auto_capture_call += 1
				solve_auto_capture = root.after((auto_capture_timer_set*60000), auto_capture)

			if(auto_capture_timer_set != 0):
				auto_capture()

	def back_click():
		onerunscreen_labelframe.destroy()

	def save_click():
		global voltage_set, m_set, s_set, m_raw, s_raw, recipient_email, folder_name, path_name
		msg = messagebox.askquestion("SAVE", "Do you want to save ?")
		if(msg=='yes'):
			if(voltage_entry.get()==''):
				messagebox.showwarning("","Please enter the voltage !")
			elif(m_entry.get()=='' or s_entry.get()==''):
				messagebox.showwarning("","Please enter the run timer !")
			elif(autocap_entry.get()==''):
				messagebox.showwarning("","Please enter the auto capture timer !")
			else:
				voltage_set = voltage_entry.get()
				m_set = int(m_entry.get())
				s_set = int(s_entry.get())
				auto_capture_timer_set = int(autocap_entry.get())

				fw = open('/home/pi/VE100/parameters1.txt','w')
				fw.writelines(str(voltage_set) + '\n')
				fw.writelines(str('%02d'%m_set) + str('%02d'%s_set) +  '\n')
				fw.writelines(str(auto_capture_timer_set) + '\n')
				fw.writelines(str(automail1_is_on) + '\n')
				messagebox.showinfo("SAVE", 'Saved !')


	back_button = Button(onerunscreen_labelframe, bd=1, bg=BUTTON_BACKGROUND_COLOR_1, font=('Courier',12, 'bold'), text='BACK', width=8, height=3, command = back_click)
	back_button.place(x=0,y=405)
	save_button = Button(onerunscreen_labelframe, bg=BUTTON_BACKGROUND_COLOR_1, font=('Courier',12, 'bold'), bd=1,text='SAVE', width=8, height=3, command = save_click)
	save_button.place(x=109,y=405)
	run1_button = Button(onerunscreen_labelframe, bg=BUTTON_BACKGROUND_COLOR_1, font=('Courier',12, 'bold'), bd=1,text='RUN', width=8, height=3, command = run1_click)
	run1_button.place(x=687,y=405)
# RUN FUNTION - END

# PREVIEW FUNTION - START
def previewScreen():
	previewscreen_labelframe = LabelFrame(root, bg='grey85', width=800, height=480)
	previewscreen_labelframe.place(x=0,y=0)
	# ~ multirunscreen0_title_labelframe = LabelFrame(multirunscreen_labelframe, text="DEFAULT STAGE", font=('Courier', 13, 'bold'), bg='green yellow', width=170, height=118)
	# ~ multirunscreen0_title_labelframe.place(x=0,y=1)
	# ~ multirunscreen1_title_labelframe = LabelFrame(multirunscreen_labelframe, text="STAGE 1", font=('Courier', 13, 'bold'), bg='grey70', width=170, height=118)
	# ~ multirunscreen1_title_labelframe.place(x=0,y=120)

	preview_labelframe = LabelFrame(previewscreen_labelframe, bg='black', fg='red', text="◉ MONITOR", font=("Courier",13,'bold'), width=622, height=410)
	preview_labelframe.place(x=173,y=0)

	limit1_canvas = Canvas(preview_labelframe, bg='grey70', bd=0, width=54, height=376)
	limit1_canvas.place(x=0,y=1)
	limit2_canvas = Canvas(preview_labelframe, bg='grey70', bd=0, width=54, height=376)
	limit2_canvas.place(x=562,y=1)
	limit_line1 = limit1_canvas.create_line(x0_limit, y0_limit, x1_limit, y1_limit, fill='red', width=2)
	limit_line2 = limit2_canvas.create_line(x0_limit, y0_limit, x1_limit, y1_limit, fill='red', width=2)
	
	button_labelframe = LabelFrame(previewscreen_labelframe, bg='grey90', width=622, height=66)
	button_labelframe.place(x=173,y=410)

	global camera
	camera_preview()
	
	def capture_clicked():
		path_name = "/home/pi/Desktop/VE100 Result/"
		image_name  = strftime("%y-%m-%d-%H-%M-%S.jpg")
		camera_capture(path_name + image_name)
		camera.stop_preview()
		
		edit_img = Image.open(path_name + image_name)
		img = ImageDraw.Draw(edit_img)
		shape = [(1024, 768), (0,550)]
		img.rectangle(shape, fill ="lightgray", outline="lightgray")
		img_font_1_17 = ImageFont.truetype("/home/pi/VE100/arial.ttf", font_size_17)
		img_font_1_26 = ImageFont.truetype("/home/pi/VE100/arial.ttf", font_size_26)
		img_font_2 = ImageFont.truetype("/home/pi/VE100/arial.ttf", 23)
		if(num_well==17):
			x_coordinate = x_coordinate_17
			for i in range(0,9):
				img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_17, fill=(0,255,0))
				x_coordinate += well_distance_17

			x_coordinate = x_coordinate - well_distance_17 + pace_17
			for i in range(9,18):
				img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_17, fill=(0,255,0))
				x_coordinate += well_distance_17
		else:
			x_coordinate = x_coordinate_26
			for i in range(0,13):
				img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_26, fill=(0,255,0))
				x_coordinate += well_distance_26

			x_coordinate = x_coordinate - well_distance_26 + pace_26
			for i in range(13,26):
				img.text((x_coordinate,y_coordinate), str(i+1), font=img_font_1_26, fill=(0,255,0))
				# ~ if(i<18):
					# ~ x_coordinate += 26
				# ~ elif(i==18):
					# ~ x_coordinate += 30
				# ~ else:
				x_coordinate += well_distance_26

		if(num_well==26):
			r1 = 568
			r2 = 568
			r3 = 568
			r4 = 568
			r5 = 568
			for i in range(0,26):
				if(i<6):
					img.text((32,r1), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
					r1 += 32
				elif(i<12):
					img.text((232,r2), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
					r2 += 32
				elif(i<18):
					img.text((432,r3), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
					r3 += 32
				elif(i<24):
					img.text((632,r4), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
					r4 += 32
				else:
					img.text((832,r5), str(i+1) + '. ' + str(well_name_list_26[i]), font=img_font_2, fill=(0,0,0))
					r5 += 32
		else:
			r1 = 568
			r2 = 568
			r3 = 568
			r4 = 568
			r5 = 568
			for i in range(0,18):
				if(i<6):
					img.text((32,r1), str(i+1) + '. ' + str(well_name_list_17[i]), font=img_font_2, fill=(0,0,0))
					r1 += 32
				elif(i<12):
					img.text((232,r2), str(i+1) + '. ' + str(well_name_list_17[i]), font=img_font_2, fill=(0,0,0))
					r2 += 32
				else:
					img.text((432,r3), str(i+1) + '. ' + str(well_name_list_17[i]), font=img_font_2, fill=(0,0,0))
					r3 += 32

		image_name  = strftime("EDIT_%y-%m-%d-%H-%M-%S-.png")
		edit_img.save(path_name + image_name,'png')

		msg = messagebox.showinfo("","Image is saved !")
		if(msg=='ok'):
			camera_preview()
			
	def back_clicked():
		camera.stop_preview()
		previewscreen_labelframe.destroy()
		main()
		
	capture_button = Button(button_labelframe, bg=BUTTON_BACKGROUND_COLOR_3, bd=1, text='CAPTURE',font=("Courier",13,'bold'), width=16, height=2, command=capture_clicked)
	capture_button.place(x=210,y=3)
	back_button = Button(button_labelframe, bg=BUTTON_BACKGROUND_COLOR_1, bd=1, text='BACK',font=("Courier",13,'bold'), width=16, height=2, command=back_clicked)
	back_button.place(x=427,y=3)
	
# PREVIEW FUNTION - END

def trial_expired():
	global trial_work_frame
	trial_work_frame = Frame(root, bg = 'white')
	trial_work_frame.pack(expand=TRUE)
	
	expire_info1_label = Label(trial_work_frame,
								text = "Your trial has expired",
								font = ('Courier',15),
								bg = 'white',
								fg = 'red')
	expire_info1_label.grid(row=0, column=0, pady=30, sticky=EW)
	
	expire_info2_label = Label(trial_work_frame,
							text = " Please enter the activation code to continue using the application",
							font = ('Courier',12),
							bg = 'white',
							fg = 'grey35')
	expire_info2_label.grid(row=2, column=0, pady=10, padx=30, sticky=W)
	
	global active_code_entry
	active_code_entry = Entry(trial_work_frame, width=30, font=('Courier',14))
	active_code_entry.grid(row=3, column=0, pady=10, padx=30, sticky=EW)
	
	activate_button = Button(trial_work_frame,
							text = "Activate",
							font = ('Helvetica', 10),
							# width = SWITCH_PAGE_BUTTON_WIDTH,
							# height = SWITCH_PAGE_BUTTON_HEIGHT,
							bg = "grey80",
							fg = "black",
							borderwidth = 0,
							command = activate_clicked)
	activate_button.grid(row=4, column=0, ipady=10, pady=30, padx=300, sticky=EW)
		
		
def trial_30days_extend():
	dt = rtc.datetime
	recent_date = dt.tm_mday
	recent_month = dt.tm_mon
	recent_year = dt.tm_year
	
	time1 = trial_year*365 + trial_month*30 + trial_date
	time2 = recent_year*365 + recent_month*30 + recent_date
	number_of_days = time2 - time1
	print("Trial days left: ", 30 - number_of_days, '/30')
	
	if(number_of_days > 30):
		trial_days = 30
		trial_expired()
		
def trial_7days():
	dt = rtc.datetime
	recent_date = dt.tm_mday
	recent_month = dt.tm_mon
	recent_year = dt.tm_year
	
	time1 = trial_year*365 + trial_month*30 + trial_date
	time2 = recent_year*365 + recent_month*30 + recent_date
	number_of_days = time2 - time1
	print("Trial days left: ", 7 - number_of_days, '/7')
	
	if(number_of_days > 7):
		trial_days = 7
		trial_expired()
		
def activate_clicked():
	active_code_enter = active_code_entry.get()
	if(active_code_enter != ''):
		if(active_code_enter == trial_30days_extend_code):
			if(active_code != trial_30days_extend_code):
				fw = open("/home/pi/VE100/active_code.txt",'w')	
				fw.writelines(active_code_enter + '\n')
				messagebox.showinfo("","Your trial package has been extended to 30 days.")
				try:
					trial_work_frame.destroy()
				except:
					pass
				main()
			else:
				messagebox.showerror("","Your code is invalid, please try again.")
		elif(active_code_enter == trial_full_active_code):
			fw = open("/home/pi/VE100/active_code.txt",'w')	
			fw.writelines(active_code_enter + '\n')
			messagebox.showinfo("","Successful activation.")
			try:
				trial_work_frame.destroy()
			except:
				pass
			main()
		else:
			messagebox.showerror("","Your code is invalid, please try again.")
	else:
		messagebox.showwarning("","Please enter activation code.")



if(active_code == trial_full_active_code):
	main()
elif(active_code == trial_30days_extend_code):
	trial_30days_extend()
else:
	trial_7days()


root.mainloop()
